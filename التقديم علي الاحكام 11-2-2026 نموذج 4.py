# ================= LIBRARIES =================
import ssl
import time
import os
import shutil
import socket
import re
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from difflib import get_close_matches
from pathlib import Path
from openpyxl import load_workbook


from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException

# webdriver_manager import kept but not used after local driver switch
from webdriver_manager.chrome import ChromeDriverManager


# ================= FIX SSL =================
ssl._create_default_https_context = ssl._create_unverified_context


# ================= GUI =================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PPO Automation - Multi Tabs")
        self.geometry("820x440")
        self.resizable(False, False)
        self.configure(bg="#eef2f7")

        self.excel_path = tk.StringVar()
        self.pdf_path = tk.StringVar()
        self.otp_code = tk.StringVar()
        self.captcha_code = tk.StringVar()
        self.attach_existing = tk.BooleanVar(value=False)
        self.debug_port = tk.StringVar(value="9222")
        self.script_dir = Path(__file__).resolve().parent
        self.status_var = tk.StringVar(value="جاهز")

        # Header
        self.header_frame = tk.Frame(self, bg="#0f172a", height=78)
        self.header_frame.pack(fill=tk.X)
        self.header_frame.pack_propagate(False)

        self.header_left = tk.Frame(self.header_frame, bg="#0f172a")
        self.header_left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.header_right = tk.Frame(self.header_frame, bg="#0f172a")
        self.header_right.pack(side=tk.RIGHT, padx=14, pady=8)

        tk.Label(
            self.header_left,
            text="Bedayti - Public Prosecution Automation System ",
            font=("Segoe UI", 16, "bold"),
            bg="#0f172a",
            fg="white"
        ).pack(anchor="w", padx=16, pady=(10, 0))
        tk.Label(
            self.header_left,
            text="تسجيل البيانات تلقائيا علي موقع النيابة العامة",
            font=("Segoe UI", 10),
            bg="#0f172a",
            fg="#cbd5e1"
        ).pack(anchor="w", padx=16, pady=(0, 8))

        tk.Label(
            self.header_right,
            text="✦ Ibrahim Saiied ✦",
            bg="#0f172a",
            fg="#fbbf24",
            font=("Segoe UI", 11, "bold")
        ).pack(anchor="e")
        tk.Label(
            self.header_right,
            text="☎ 01144366104       ",
            bg="#0f172a",
            fg="#e2e8f0",
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="e")

        # Main card
        self.main_card = tk.Frame(self, bg="white", bd=1, relief="solid", padx=14, pady=12)
        self.main_card.pack(fill=tk.BOTH, expand=True, padx=14, pady=12)

        # إطار اختيار الملفات
        self.files_frame = tk.LabelFrame(
            self.main_card, text="الملفات", bg="white", fg="#0f172a", font=("Segoe UI", 10, "bold"), padx=10, pady=8
        )
        self.files_frame.pack(fill=tk.X, pady=(0, 8))

        row1 = tk.Frame(self.files_frame, bg="white")
        row1.pack(fill=tk.X, pady=4)
        tk.Label(row1, text="ملف Excel", bg="white").pack(side=tk.LEFT, padx=(0, 8))
        tk.Entry(row1, textvariable=self.excel_path, width=75, relief="solid", bd=1).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(row1, text="اختيار Excel", command=self.select_excel, bg="#e2e8f0").pack(side=tk.LEFT)

        row2 = tk.Frame(self.files_frame, bg="white")
        row2.pack(fill=tk.X, pady=4)
        tk.Label(row2, text="ملف PDF (التوكيل)", bg="white").pack(side=tk.LEFT, padx=(0, 8))
        tk.Entry(row2, textvariable=self.pdf_path, width=68, relief="solid", bd=1).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(row2, text="اختيار PDF", command=self.select_pdf, bg="#e2e8f0").pack(side=tk.LEFT)

        # وضع التجربة على متصفح مفتوح
        self.attach_frame = tk.LabelFrame(
            self.main_card, text="الاتصال بالمتصفح", bg="white", fg="#0f172a", font=("Segoe UI", 10, "bold"), padx=10, pady=8
        )
        self.attach_frame.pack(fill=tk.X, pady=(0, 8))
        tk.Checkbutton(
            self.attach_frame,
            text="تجربة على متصفح مفتوح (تخطي تسجيل الدخول)",
            variable=self.attach_existing,
            bg="white"
        ).pack(side=tk.LEFT, padx=4)
        tk.Label(self.attach_frame, text="Port", bg="white").pack(side=tk.LEFT, padx=(12, 4))
        tk.Entry(self.attach_frame, textvariable=self.debug_port, width=10, relief="solid", bd=1).pack(side=tk.LEFT, padx=4)

        # شريط الأوامر
        self.actions_frame = tk.Frame(self.main_card, bg="white")
        self.actions_frame.pack(fill=tk.X, pady=(2, 10))
        self.start_btn = tk.Button(
            self.actions_frame,
            text="يلا بينا",
            command=self.start,
            bg="#16a34a",
            fg="white",
            activebackground="#15803d",
            activeforeground="white",
            font=("Segoe UI", 10, "bold"),
            relief="raised",
            bd=2,
            padx=19,
            pady=4
        )
        self.start_btn.pack(anchor="center")

        # زر إعادة التحميل اليدوي للتعافي من الأخطاء أو عدم تحميل الكابتشا
        self.reload_btn = tk.Button(
            self.actions_frame,
            text="إعادة تحميل الطلب الحالي",
            command=self.reload_current,
            state=tk.DISABLED,
            bg="#f59e0b",
            fg="white",
            activebackground="#d97706",
            activeforeground="white",
            padx=12
        )

        # إطار OTP (مخفي حتى نحتاجه)
        self.otp_frame = tk.LabelFrame(
            self.main_card, text="OTP", bg="white", fg="#0f172a", font=("Segoe UI", 10, "bold"), padx=10, pady=8
        )
        tk.Label(self.otp_frame, text="OTP", bg="white").pack(side=tk.LEFT)
        tk.Entry(self.otp_frame, textvariable=self.otp_code, width=24, relief="solid", bd=1).pack(side=tk.LEFT, padx=6)
        tk.Button(self.otp_frame, text="تأكيد", command=self.submit_otp, bg="#10b981", fg="white").pack(side=tk.LEFT)

        # التحكم اليدوي في الكابتشا والخطوات (مخفي حتى نحتاجه)
        self.step_frame = tk.LabelFrame(
            self.main_card, text="الكابتشا", bg="white", fg="#0f172a", font=("Segoe UI", 10, "bold"), padx=10, pady=8
        )
        tk.Label(self.step_frame, text="ادخل الكابتشا:", bg="white").pack(side=tk.LEFT)
        tk.Entry(self.step_frame, textvariable=self.captcha_code, width=22, relief="solid", bd=1).pack(side=tk.LEFT, padx=6)
        self.next_btn = tk.Button(
            self.step_frame,
            text="الخطوة التالية",
            command=self.next_step,
            state=tk.DISABLED,
            bg="#2563eb",
            fg="white",
            activebackground="#1d4ed8",
            activeforeground="white"
        )
        self.next_btn.pack(side=tk.LEFT)

        # حالة التنفيذ
        self.status_frame = tk.Frame(self.main_card, bg="#f8fafc", bd=1, relief="solid")
        self.status_frame.pack(fill=tk.X, pady=(8, 0))
        tk.Label(
            self.status_frame,
            textvariable=self.status_var,
            fg="#0f172a",
            bg="#f8fafc",
            anchor="w",
            padx=10,
            pady=8
        ).pack(fill=tk.X)

        self.driver = None
        self.wait = None
        self.fixed = {}
        self.cases = None
        self.current_index = 0
        self.case_tabs = []
        self.state = "idle"  # idle -> waiting_captcha -> ready
        self.set_default_file_paths()

    # ================= HELPERS =================
    def err(self, msg, raise_exc=True):
        messagebox.showerror("خطأ", msg)
        if raise_exc:
            raise Exception(msg)

    def hide_prestart_widgets(self):
        for w in (self.files_frame, self.attach_frame):
            try:
                w.pack_forget()
            except Exception:
                pass
        try:
            self.start_btn.config(state=tk.DISABLED)
            self.start_btn.pack_forget()
        except Exception:
            pass

    def show_reload_button(self):
        try:
            if not self.reload_btn.winfo_ismapped():
                self.reload_btn.pack(side=tk.LEFT, padx=(8, 0))
            self.reload_btn.config(state=tk.NORMAL)
        except Exception:
            pass

    def hide_runtime_widgets_on_finish(self):
        try:
            self.step_frame.pack_forget()
        except Exception:
            pass
        try:
            self.otp_frame.pack_forget()
        except Exception:
            pass
        try:
            self.reload_btn.config(state=tk.DISABLED)
            self.reload_btn.pack_forget()
        except Exception:
            pass

    def build_driver(self, options):
        # 1) Prefer local chromedriver next to script (offline-safe)
        local_driver = self.script_dir / "chromedriver.exe"
        if local_driver.is_file():
            service = Service(str(local_driver))
            return webdriver.Chrome(service=service, options=options)

        # 2) Try Selenium Manager (uses PATH / auto resolution)
        try:
            return webdriver.Chrome(options=options)
        except Exception:
            pass

        # 3) Fallback to webdriver_manager with SSL verify disabled
        # Useful behind company proxy/self-signed cert chains.
        os.environ.setdefault("WDM_SSL_VERIFY", "0")
        service = Service(ChromeDriverManager().install())
        return webdriver.Chrome(service=service, options=options)

    def load_input_data(self):
        if not self.excel_path.get() or not self.pdf_path.get():
            self.err("اختار ملف Excel و PDF")

        if not os.path.isfile(self.excel_path.get()):
            self.err("مسار ملف Excel غير موجود")
        if not os.path.isfile(self.pdf_path.get()):
            self.err("مسار ملف PDF غير موجود")

        # Read Excel
        df_fixed = pd.read_excel(self.excel_path.get(), sheet_name="Fixed_Data")
        self.fixed = dict(zip(df_fixed["الحقل"], df_fixed["البيانات"]))

        self.cases = pd.read_excel(self.excel_path.get(), sheet_name="Cases_Data")
        if self.cases.empty:
            self.err("Cases_Data فاضي")

    def start_from_open_browser(self):
        port = (self.debug_port.get() or "").strip()
        if not port.isdigit():
            self.err("رقم الـ Port غير صحيح. مثال: 9222")

        port_num = int(port)
        if not self.is_debug_port_open(port_num):
            self.err(
                "لم يتم العثور على Chrome debug على هذا الـ Port.\n"
                "شغّل كروم بهذا الأمر أولًا:\n"
                "chrome.exe --remote-debugging-port=9222 --user-data-dir=C:\\chrome-debug"
            )

        self.status_var.set("جاري الاتصال بالمتصفح المفتوح...")
        options = Options()
        options.add_experimental_option("debuggerAddress", f"127.0.0.1:{port}")
        self.driver = self.build_attach_driver(options)
        self.wait = WebDriverWait(self.driver, 30)

        # ابحث في كل التبويبات/الإطارات عن صفحة الطلبات (P23_CAUSE_NUMBER)
        # لا توقف التنفيذ مبكرًا: أحيانًا الحقل يتأخر أو يكون داخل iframe.
        found_requests = self.switch_to_requests_tab(timeout_per_tab=6)

        try:
            self.otp_frame.pack_forget()
        except Exception:
            pass
        self.hide_prestart_widgets()
        self.show_reload_button()
        self.current_index = 0
        self.case_tabs = []
        if found_requests:
            self.status_var.set("متصل بمتصفح مفتوح. تم العثور على صفحة الطلبات.")
        else:
            self.status_var.set("متصل بمتصفح مفتوح. لم يتم تأكيد الصفحة وسيتم المحاولة مباشرة.")
        self.prepare_all_cases_tabs()

    def is_debug_port_open(self, port, host="127.0.0.1", timeout=1.5):
        try:
            with socket.create_connection((host, port), timeout=timeout):
                return True
        except OSError:
            return False

    def find_local_chromedriver(self):
        local_driver = self.script_dir / "chromedriver.exe"
        if local_driver.is_file():
            return str(local_driver)

        on_path = shutil.which("chromedriver")
        if on_path:
            return on_path

        return None

    def build_attach_driver(self, options):
        """
        Driver creation for "attach to open browser":
        1) local chromedriver next to script
        2) chromedriver on PATH
        3) Selenium Manager
        """
        driver_path = self.find_local_chromedriver()
        if driver_path:
            return webdriver.Chrome(service=Service(driver_path), options=options)

        try:
            return webdriver.Chrome(options=options)
        except Exception as e:
            self.err(
                "تعذر الاتصال بالمتصفح المفتوح.\n"
                "ثبّت chromedriver (بجانب السكربت أو على PATH) أو تأكد من توافق Chrome/Selenium.\n"
                f"التفاصيل: {e}"
            )

    def switch_to_requests_tab(self, timeout_per_tab=6):
        """
        جرّب كل تبويبات المتصفح حتى نجد صفحة الطلبات التي تحتوي P23_CAUSE_NUMBER.
        """
        d = self.driver
        handles = d.window_handles
        for handle in handles:
            try:
                d.switch_to.window(handle)
                d.switch_to.default_content()
                WebDriverWait(d, timeout_per_tab).until(
                    EC.presence_of_element_located((By.NAME, "P23_CAUSE_NUMBER"))
                )
                return True
            except Exception:
                # أحيانًا الصفحة تكون داخل iframe؛ جرّب البحث داخل الإطارات
                try:
                    d.switch_to.default_content()
                    iframes = d.find_elements(By.TAG_NAME, "iframe")
                    for frm in iframes:
                        try:
                            d.switch_to.default_content()
                            d.switch_to.frame(frm)
                            WebDriverWait(d, 1.5).until(
                                EC.presence_of_element_located((By.NAME, "P23_CAUSE_NUMBER"))
                            )
                            return True
                        except Exception:
                            pass
                except Exception:
                    pass
                finally:
                    try:
                        d.switch_to.default_content()
                    except Exception:
                        pass
                continue
        return False

    def get_fixed(self, key):
        if key not in self.fixed or pd.isna(self.fixed[key]):
            self.err(f"الحقل [{key}] ناقص في Fixed_Data")
        return str(self.fixed[key]).strip()

    def get_case(self, row, col):
        if col not in row or pd.isna(row[col]):
            self.err(f"الحقل [{col}] ناقص في Cases_Data\nصف {row.name+2}")
        return str(row[col]).strip()

    def clear_and_type(self, locator_by, locator_value, text):
        e = self.wait.until(EC.presence_of_element_located((locator_by, locator_value)))
        try:
            e.clear()
        except Exception:
            pass
        e.send_keys(text)

    def select_option_fuzzy(self, locator_by, locator_value, text):
        """اختر من dropdown بأقرب مطابقة (مرن في التطابق)"""
        select = Select(self.wait.until(EC.presence_of_element_located((locator_by, locator_value))))
        
        # احصل على كل الخيارات
        all_options = [option.text.strip() for option in select.options]
        
        # ابحث عن تطابق دقيق أولاً
        if text in all_options:
            select.select_by_visible_text(text)
            return
        
        # ابحث عن أقرب خيار
        matches = get_close_matches(text, all_options, n=1, cutoff=0.6)
        if matches:
            select.select_by_visible_text(matches[0])
            return
        
        # إذا لم يوجد تطابق، رفع خطأ مع إظهار الخيارات المتاحة
        self.err(f"لم يُعثر على '{text}' في الخيارات المتاحة:\n{', '.join(all_options)}")

    def wait_dropdown_loaded(self, locator_by, locator_value, expected_text=None, timeout=12):
        """
        انتظر حتى تحميل خيارات الـ dropdown التابعة (مثل قسم الشرطة بعد اختيار المحافظة).
        إذا تم تمرير expected_text ينتظر حتى يظهر (أو أقرب مطابقة) داخل الخيارات.
        """
        def _is_loaded(_):
            try:
                elem = self.driver.find_element(locator_by, locator_value)
                opts = [o.text.strip() for o in Select(elem).options if o.text.strip()]
                if not opts:
                    return False

                # تجاهل خيار placeholder الشائع
                normalized = [o for o in opts if o not in ("اختر", "اختر...", "-- اختر --")]
                if not normalized:
                    return False

                if expected_text:
                    if expected_text in normalized:
                        return True
                    return bool(get_close_matches(expected_text, normalized, n=1, cutoff=0.6))

                # بدون نص متوقع: يكفي وجود أكثر من خيار فعلي
                return len(normalized) >= 2
            except Exception:
                return False

        WebDriverWait(self.driver, timeout).until(_is_loaded)

    def switch_to_dialog_frame(self, frame_css, field_id, timeout=12):
        """
        انقل للـ iframe الخاص بالنافذة المنبثقة.
        يحاول أولًا عبر frame_css، ثم fallback بالبحث عن iframe يحتوي field_id.
        """
        d = self.driver
        try:
            d.switch_to.default_content()
        except Exception:
            pass

        try:
            WebDriverWait(d, timeout).until(
                EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR, frame_css))
            )
            WebDriverWait(d, timeout).until(EC.presence_of_element_located((By.ID, field_id)))
            return
        except Exception:
            d.switch_to.default_content()

        # fallback: جرّب كل iframes حتى نجد الحقل المطلوب
        end_time = time.time() + timeout
        while time.time() < end_time:
            iframes = d.find_elements(By.TAG_NAME, "iframe")
            for frm in iframes:
                try:
                    d.switch_to.default_content()
                    d.switch_to.frame(frm)
                    if d.find_elements(By.ID, field_id):
                        return
                except Exception:
                    pass
            time.sleep(0.2)
        d.switch_to.default_content()
        raise TimeoutException(f"لم يتم العثور على iframe يحتوي الحقل {field_id}")

    def find_latest_file(self, patterns):
        candidates = []
        for pattern in patterns:
            candidates.extend(self.script_dir.glob(pattern))

        files = [f for f in candidates if f.is_file()]
        if not files:
            return ""

        latest = max(files, key=lambda f: f.stat().st_mtime)
        return str(latest.resolve())

    def set_default_file_paths(self):
        excel_default = self.find_latest_file(["*.xlsx", "*.xls"])
        pdf_default = self.find_latest_file(["*.pdf"])

        if excel_default:
            self.excel_path.set(excel_default)
        if pdf_default:
            self.pdf_path.set(pdf_default)

    # ================= FILE PICKERS =================
    def select_excel(self):
        self.excel_path.set(filedialog.askopenfilename(filetypes=[["Excel", "*.xlsx;*.xls"]]))

    def select_pdf(self):
        self.pdf_path.set(filedialog.askopenfilename(filetypes=[["PDF", "*.pdf"]]))

    # ================= START =================
    def start(self):
        try:
            self.load_input_data()

            if self.attach_existing.get():
                self.start_from_open_browser()
                return

            # Chrome using local chromedriver
            options = Options()
            options.add_argument("--start-maximized")
            options.add_experimental_option("detach", True)
            port = (self.debug_port.get() or "9222").strip()
            if not port.isdigit():
                port = "9222"
            options.add_argument(f"--remote-debugging-port={port}")

            self.driver = self.build_driver(options)

            self.wait = WebDriverWait(self.driver, 30)

            self.driver.get("https://www.ppo.gov.eg/ppo/r/ppoportal/ppoportal/login-page")

            self.wait.until(EC.presence_of_element_located((By.NAME, "P9999_USERNAME"))).send_keys(self.get_fixed("اسم_المستخدم"))
            self.wait.until(EC.presence_of_element_located((By.NAME, "P9999_PASSWORD"))).send_keys(self.get_fixed("الرقم_السري"))
            self.wait.until(EC.element_to_be_clickable((By.ID, "GENERATE_OTP"))).click()

            # أخفِ عناصر الإعداد بعد بدء التشغيل
            self.hide_prestart_widgets()

            # أظهر إطار OTP فقط
            self.otp_frame.pack(fill=tk.X, pady=8)
            self.status_var.set("انتظر إدخال OTP")

        except Exception as e:
            self.err(f"فشل البدء: {e}", raise_exc=False)

    # ================= OTP =================
    def submit_otp(self):
        try:
            self.wait.until(EC.presence_of_element_located((By.ID, "P9999_OTP_VER"))).send_keys(self.otp_code.get())
            self.wait.until(EC.element_to_be_clickable((By.ID, "LOGIN"))).click()

            # تأكد من الانتقال
            self.wait.until(EC.presence_of_element_located((By.ID, "navbarDropdownMenuLink")))
            self.after_login()

        except Exception as e:
            self.err(f"فشل OTP: {e}", raise_exc=False)

    # ================= AFTER LOGIN =================
    def after_login(self):
        d, w = self.driver, self.wait
        try:
            # Open cases page
            w.until(EC.element_to_be_clickable((By.ID, "navbarDropdownMenuLink"))).click()
            w.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".dropdown-menu .dropdown-item:nth-child(2)"))).click()
            w.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.a-CardView-fullLink[onclick*=\"go_to_page('1244','23')\"]"))).click()

            w.until(EC.presence_of_element_located((By.NAME, "P23_CAUSE_NUMBER")))
            self.status_var.set("تم فتح صفحة الطلبات")

            # أخفِ OTP عند الانتهاء منه
            try:
                self.otp_frame.pack_forget()
            except Exception:
                pass

            # تفعيل زر إعادة التحميل بعد الدخول
            self.show_reload_button()

            # تجهيز كل الطلبات في تبويبات ثم الوقوف على كابتشا أول طلب
            self.current_index = 0
            self.case_tabs = []
            self.prepare_all_cases_tabs()

        except Exception as e:
            self.err(f"فشل بعد تسجيل الدخول: {e}", raise_exc=False)

    # ================= MULTI-TAB CASE PREPARATION =================
    def prepare_case_in_current_tab(self, row, idx, total):
        d, w = self.driver, self.wait
        self.status_var.set(f"تجهيز الطلب {idx+1}/{total} | رقم الطلب: {row['رقم_الطلب']}")

        # ===== CASE =====
        self.clear_and_type(By.NAME, "P23_CAUSE_NUMBER", self.get_case(row, "رقم_القضية"))
        self.clear_and_type(By.NAME, "P23_CAUSE_YEAR", self.get_case(row, "سنة_القضية"))
        self.select_option_fuzzy(By.ID, "P23_TABLE", self.get_case(row, "الجدول"))
        self.select_option_fuzzy(By.ID, "P23_GOV", self.get_case(row, "المحافظة"))
        police_department = self.get_case(row, "قسم_الشرطة")
        self.wait_dropdown_loaded(By.ID, "P23_POLICE_DEPARTMENT", expected_text=police_department, timeout=12)
        self.select_option_fuzzy(By.ID, "P23_POLICE_DEPARTMENT", police_department)
        self.select_option_fuzzy(By.ID, "P23_SEND_TO", self.get_case(row, "الي"))

        # ===== FIXED =====
        self.select_option_fuzzy(By.ID, "P23_AGENT_DESCRIPTION", self.get_fixed("توصيف_الوكيل"))
        self.clear_and_type(By.ID, "P23_AGENT_NUMBER", self.get_fixed("رقم_التوكيل"))
        self.clear_and_type(By.ID, "P23_CARD_NUMBER", self.get_fixed("رقم_الكارنية"))
        self.select_option_fuzzy(By.ID, "P23_ENTRY_TYPE", self.get_fixed("نوع_القيد"))
        self.clear_and_type(By.ID, "P23_ENTITY", self.get_fixed("جهة_إصدار_التوكيل"))

        # ===== ADD CLIENT =====
        w.until(EC.element_to_be_clickable((By.ID, "B1"))).click()
        iframe = w.until(EC.presence_of_element_located((By.TAG_NAME, "iframe")))
        d.switch_to.frame(iframe)

        self.select_option_fuzzy(By.ID, "P26_PETITIONER_DESC", self.get_fixed("صفة_مقدم_الطلب"))
        self.select_option_fuzzy(By.ID, "P26_IDENTITY_TYPE", self.get_fixed("نوع_الهوية"))

        self.clear_and_type(By.ID, "P26_NATIONAL_ID", self.get_fixed("الرقم_القومي"))
        self.clear_and_type(By.ID, "P26_FIRST_NAME", self.get_fixed("الاسم_الاول"))
        self.clear_and_type(By.ID, "P26_SECOND_NAME", self.get_fixed("الاسم_الثاني"))
        self.clear_and_type(By.ID, "P26_THIRD_NAME", self.get_fixed("الاسم_الثالث"))
        self.clear_and_type(By.ID, "P26_FOURTH_NAME", self.get_fixed("الاسم_الرابع"))

        self.clear_and_type(By.ID, "P26_ADDRESS", self.get_fixed("العنوان"))
        self.clear_and_type(By.ID, "P26_EMAIL", self.get_fixed("البريد_الالكتروني"))

        w.until(EC.element_to_be_clickable((By.ID, "B3"))).click()
        d.switch_to.default_content()

        # ===== ADD DOCUMENT =====
        doc_btn_locator = (
            By.CSS_SELECTOR,
            "button[aria-label='إضافة مستند'], "
            "button[title='إضافة مستند'], "
            "button[onclick*='add-attatchment']"
        )
        w.until(EC.element_to_be_clickable(doc_btn_locator)).click()
        self.switch_to_dialog_frame("iframe[src*='add-attatchment']", "P21_ATTATCHMENT_TYPE", timeout=15)

        self.select_option_fuzzy(By.ID, "P21_ATTATCHMENT_TYPE", self.get_fixed("نوع_المستند"))
        w.until(EC.presence_of_element_located((By.ID, "P21_ATTATCHMENT_input"))).send_keys(os.path.abspath(self.pdf_path.get()))
        w.until(EC.element_to_be_clickable((By.ID, "B3"))).click()
        d.switch_to.default_content()
        WebDriverWait(d, 12).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, "iframe[src*='add-attatchment']"))
        )
        w.until(EC.element_to_be_clickable((By.ID, "P23_RECEIPT")))
        time.sleep(0.5)

        # ===== DELIVERY =====
        self.select_option_fuzzy(By.ID, "P23_RECEIPT", self.get_fixed("طريقة_الإستلام"))
        self.select_option_fuzzy(By.ID, "P23_DELIVERY_GOV", self.get_fixed("محافظة_التوصيل"))
        self.clear_and_type(By.ID, "P23_CONTACT_PHONE_NUMBER", self.get_fixed("رقم_تليفون_للتواصل"))
        self.clear_and_type(By.ID, "P23_DELIVERY_ADD", self.get_fixed("عنوان_التوصيل"))
        w.until(EC.element_to_be_clickable((By.ID, "P23_TERMS_CONDITIONS_LABEL"))).click()

        d.switch_to.default_content()

    def open_request_tab(self, request_url):
        d, w = self.driver, self.wait
        d.execute_script("window.open(arguments[0], '_blank');", request_url)
        d.switch_to.window(d.window_handles[-1])
        d.switch_to.default_content()
        w.until(EC.presence_of_element_located((By.NAME, "P23_CAUSE_NUMBER")))

    def switch_to_case_tab(self, index):
        if index < 0 or index >= len(self.case_tabs):
            self.err("فهرس التبويب المطلوب غير صحيح")
        self.driver.switch_to.window(self.case_tabs[index])
        self.driver.switch_to.default_content()

    def focus_captcha_field(self):
        d = self.driver
        try:
            try:
                cap_elem = WebDriverWait(d, 2).until(
                    EC.presence_of_element_located((By.ID, "P23_CAPTCHA_CODE"))
                )
            except Exception:
                cap_elem = WebDriverWait(d, 4).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[name*='captcha'], input[id*='captcha']"))
                )
            d.execute_script("arguments[0].scrollIntoView({block:'center'});", cap_elem)
            try:
                cap_elem.click()
            except Exception:
                pass
        except Exception:
            pass

    def fill_captcha_in_current_tab(self, captcha_text):
        w = self.wait
        cap_elem = w.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input[name*='captcha'], input[id*='captcha']"))
        )
        try:
            cap_elem.clear()
        except Exception:
            pass
        cap_elem.send_keys(captcha_text)

    def click_first_visible(self, locators):
        d = self.driver
        for by, locator in locators:
            try:
                elements = d.find_elements(by, locator)
            except Exception:
                continue
            for elem in elements:
                try:
                    if not elem.is_displayed():
                        continue
                    d.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
                    try:
                        elem.click()
                    except Exception:
                        d.execute_script("arguments[0].click();", elem)
                    return True
                except Exception:
                    continue
        return False

    def click_submit_request_button(self):
        # ترتيب محاولات مرن لأن IDs في APEX غالبًا متغيرة
        submit_locators = [
            (By.ID, "SUBMIT"),
            (By.CSS_SELECTOR, "button[id*='SUBMIT'],input[id*='SUBMIT'][type='button'],input[id*='SUBMIT'][type='submit']"),
            (By.CSS_SELECTOR, "button[title*='تقديم'],button[aria-label*='تقديم'],button[title*='إرسال'],button[aria-label*='إرسال']"),
            (By.CSS_SELECTOR, "button[onclick*='apex.submit'],a[onclick*='apex.submit'],button[onclick*='submit'],a[onclick*='submit']"),
            (By.XPATH, "//button[contains(normalize-space(.),'تقديم') or contains(normalize-space(.),'إرسال') or contains(normalize-space(.),'ارسال') or contains(normalize-space(.),'Submit')]"),
            (By.XPATH, "//a[contains(normalize-space(.),'تقديم') or contains(normalize-space(.),'إرسال') or contains(normalize-space(.),'ارسال') or contains(normalize-space(.),'Submit')]"),
        ]
        return self.click_first_visible(submit_locators)

    def click_optional_confirm(self):
        confirm_locators = [
            (By.XPATH, "//button[contains(normalize-space(.),'نعم') or contains(normalize-space(.),'تأكيد') or contains(normalize-space(.),'موافق') or normalize-space(.)='OK']"),
            (By.XPATH, "//a[contains(normalize-space(.),'نعم') or contains(normalize-space(.),'تأكيد') or contains(normalize-space(.),'موافق') or normalize-space(.)='OK']"),
            (By.CSS_SELECTOR, "button.ui-button--hot, .ui-dialog-buttonset button.t-Button--hot"),
        ]
        return self.click_first_visible(confirm_locators)

    def detect_submission_error(self):
        d = self.driver
        selectors = [
            ".t-Alert--danger",
            ".a-Notification--error",
            ".u-danger-text",
            ".t-Form-error",
        ]
        text = ""
        for sel in selectors:
            try:
                elems = d.find_elements(By.CSS_SELECTOR, sel)
                for e in elems:
                    if e.is_displayed():
                        txt = e.text.strip()
                        if txt:
                            text += f" {txt}"
            except Exception:
                pass

        text = text.strip()
        if not text:
            return None

        # اعتبرها رسالة خطأ فقط إذا كانت دلالتها واضحة
        lowered = text.lower()
        keywords = ["خطأ", "غير صحيح", "كود التحقق", "captcha", "تحقق", "invalid"]
        if any(k in lowered for k in keywords):
            return text
        return None

    def submit_current_request(self, captcha_text):
        self.fill_captcha_in_current_tab(captcha_text)

        if not self.click_submit_request_button():
            return False, "تعذر العثور على زر تقديم الطلب."

        # أحيانًا يظهر تأكيد إضافي بعد الضغط على تقديم
        time.sleep(0.3)
        self.click_optional_confirm()

        # امنح الصفحة فرصة لإظهار نتيجة التقديم
        time.sleep(1.0)
        err_text = self.detect_submission_error()
        if err_text:
            return False, err_text
        return True, ""

    def click_series_refresh(self):
        refresh_locators = [
            (By.XPATH, "//button[.//i[contains(@class,'cc-refresh')]]"),
            (By.XPATH, "//a[.//i[contains(@class,'cc-refresh')]]"),
            (By.CSS_SELECTOR, "i.cc-refresh"),
            (By.CSS_SELECTOR, ".cc-refresh"),
        ]
        return self.click_first_visible(refresh_locators)

    def read_series_display_text(self):
        d = self.driver
        elem = WebDriverWait(d, 4).until(
            EC.presence_of_element_located((By.ID, "P40_SERIES_DISPLAY"))
        )
        txt = (elem.text or "").strip()
        if not txt:
            txt = (elem.get_attribute("value") or "").strip()
        return txt

    def fetch_request_number_current_tab(self, max_refresh_clicks=5):
        d = self.driver
        d.switch_to.default_content()

        for _ in range(max_refresh_clicks):
            self.click_series_refresh()
            time.sleep(0.7)
            try:
                txt = self.read_series_display_text()
            except Exception:
                txt = ""

            # عندما يتحول النص لرقم طلب
            if txt and re.fullmatch(r"\d+", txt):
                return txt

            # ما زال في حالة الإنشاء أو نص غير رقمي
            if txt == "جارى إنشاء الطلب.":
                continue

        return ""

    def save_request_numbers_to_excel(self):
        path = self.excel_path.get()
        if not path or not os.path.isfile(path):
            self.err("مسار ملف Excel غير موجود لحفظ أرقام الطلبات.")

        try:
            wb = load_workbook(path)
            if "Cases_Data" not in wb.sheetnames:
                self.err("لم يتم العثور على الشيت Cases_Data في ملف Excel.")
            ws = wb["Cases_Data"]

            # ابحث عن عمود رقم_الطلب في رأس الجدول
            headers = {}
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=c).value
                if val is not None:
                    headers[str(val).strip()] = c

            req_col = headers.get("رقم_الطلب")
            if not req_col:
                req_col = ws.max_column + 1
                ws.cell(row=1, column=req_col, value="رقم_الطلب")

            # اكتب الأرقام بنفس ترتيب الصفوف في Cases_Data
            for idx in range(len(self.cases)):
                v = self.cases.iloc[idx].get("رقم_الطلب", "")
                if pd.isna(v):
                    v = ""
                ws.cell(row=idx + 2, column=req_col, value=str(v).strip())

            wb.save(path)
        except PermissionError:
            self.err("تعذر حفظ ملف Excel. أغلق الملف من Excel ثم أعد المحاولة.")

    def collect_request_numbers_all_tabs(self):
        total = len(self.case_tabs)
        if total == 0:
            return

        if "رقم_الطلب" not in self.cases.columns:
            self.cases["رقم_الطلب"] = ""

        for idx in range(total):
            self.switch_to_case_tab(idx)
            self.status_var.set(f"جمع رقم الطلب {idx+1}/{total} ...")
            req_no = self.fetch_request_number_current_tab(max_refresh_clicks=5)
            self.cases.at[self.cases.index[idx], "رقم_الطلب"] = req_no if req_no else ""

        self.save_request_numbers_to_excel()

    def activate_current_case_for_captcha(self):
        self.switch_to_case_tab(self.current_index)
        self.focus_captcha_field()
        if not self.step_frame.winfo_ismapped():
            self.step_frame.pack(fill=tk.X, pady=8)
        self.next_btn.config(state=tk.NORMAL)
        self.state = "waiting_captcha"
        row = self.cases.iloc[self.current_index]
        self.status_var.set(
            f"جاهز للكابتشا {self.current_index+1}/{len(self.cases)} | رقم الطلب: {row['رقم_الطلب']}"
        )

    def prepare_all_cases_tabs(self):
        d, w = self.driver, self.wait
        try:
            total = len(self.cases)
            if total == 0:
                self.err("Cases_Data فاضي")

            try:
                self.step_frame.pack_forget()
            except Exception:
                pass
            self.next_btn.config(state=tk.DISABLED)
            self.state = "ready"
            self.case_tabs = []
            self.current_index = 0

            d.switch_to.default_content()
            w.until(EC.presence_of_element_located((By.NAME, "P23_CAUSE_NUMBER")))
            request_url = d.current_url

            for idx in range(total):
                if idx == 0:
                    d.switch_to.default_content()
                else:
                    self.open_request_tab(request_url)

                row = self.cases.iloc[idx]
                self.prepare_case_in_current_tab(row, idx, total)
                self.case_tabs.append(d.current_window_handle)

            self.current_index = 0
            self.activate_current_case_for_captcha()
            messagebox.showinfo("جاهز", f"تم تجهيز {total} طلب في تبويبات منفصلة.\nابدأ بإدخال كابتشا الطلب الأول.")
        except Exception as e:
            self.err(f"فشل تجهيز الطلبات في التبويبات: {e}", raise_exc=False)

    # ================= MANUAL RECOVERY / RELOAD =================
    def reload_current(self):
        try:
            if not self.driver or not self.wait:
                self.err("المتصفح غير مُهيأ بعد. ابدأ التشغيل أولاً.")
            if not self.case_tabs:
                self.err("لا يوجد طلبات مجهزة لإعادة تحميلها بعد.")

            self.status_var.set("جاري إعادة التحميل...")
            self.switch_to_case_tab(self.current_index)

            # ارجع للإطار الأساسي ثم أعد تحميل الصفحة
            try:
                self.driver.switch_to.default_content()
            except Exception:
                pass

            self.driver.refresh()

            # انتظر ظهور نموذج الطلبات
            self.wait.until(EC.presence_of_element_located((By.NAME, "P23_CAUSE_NUMBER")))

            # أخفِ إطار الكابتشا إن كان ظاهرًا
            try:
                self.step_frame.pack_forget()
            except Exception:
                pass

            row = self.cases.iloc[self.current_index]
            self.prepare_case_in_current_tab(row, self.current_index, len(self.cases))
            self.activate_current_case_for_captcha()
            self.status_var.set("تمت إعادة تجهيز الطلب الحالي. أدخل الكابتشا مجددًا.")
        except Exception as e:
            self.err(f"فشل إعادة التحميل: {e}", raise_exc=False)

    # استكمال بعد إدخال الكابتشا والضغط على الخطوة التالية
    def next_step(self):
        w = self.wait
        try:
            if self.state != "waiting_captcha":
                return
            if not self.case_tabs:
                self.err("لا يوجد تبويبات طلبات للتنقل بينها.")

            self.switch_to_case_tab(self.current_index)
            captcha_value = (self.captcha_code.get() or "").strip()
            if not captcha_value:
                self.err("ادخل الكابتشا أولًا.")

            # اكتب الكابتشا في الحقل المحدد ثم اضغط زر التقديم المحدد
            captcha_elem = w.until(EC.presence_of_element_located((By.ID, "P23_CAPTCHA_CODE")))
            try:
                captcha_elem.clear()
            except Exception:
                pass
            captcha_elem.send_keys(captcha_value)

            submit_btn = w.until(EC.element_to_be_clickable((By.ID, "cid-submit")))
            submit_btn.click()
            time.sleep(0.6)

            # إخفاء إطار الكابتشا مؤقتاً
            try:
                self.step_frame.pack_forget()
            except Exception:
                pass

            # تم تقديم الطلب الحالي، انتقل للطلب التالي الجاهز
            self.captcha_code.set("")
            self.state = "ready"
            self.next_btn.config(state=tk.DISABLED)

            if self.current_index + 1 >= len(self.case_tabs):
                self.status_var.set("تم تقديم كل الطلبات. جارٍ جمع أرقام الطلبات...")
                self.collect_request_numbers_all_tabs()
                self.hide_runtime_widgets_on_finish()
                self.status_var.set("تم حفظ أرقام الطلبات في Excel.")
                messagebox.showinfo("انتهى", "تم تقديم جميع الطلبات وحفظ أرقام الطلبات في عمود رقم_الطلب.")
                return

            self.current_index += 1
            self.activate_current_case_for_captcha()

        except Exception as e:
            self.err(f"فشل الاستكمال بعد الكابتشا: {e}", raise_exc=False)


# ================= RUN =================
if __name__ == "__main__":
    App().mainloop()

