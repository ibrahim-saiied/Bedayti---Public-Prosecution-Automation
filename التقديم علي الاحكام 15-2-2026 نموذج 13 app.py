# ================= LIBRARIES =================
import ssl
import time
import os
import shutil
import re
import io
import sys
import json
import hashlib
import subprocess
import threading
import urllib.request
import urllib.error
import webbrowser
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from difflib import get_close_matches
from pathlib import Path
from openpyxl import load_workbook
try:
    from PIL import Image, ImageOps
except Exception:
    Image = None
    ImageOps = None

try:
    import pytesseract
except Exception:
    pytesseract = None


from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException

from webdriver_manager.chrome import ChromeDriverManager
from license_service import (
    get_runtime_dir,
    read_saved_license_key,
    save_license_key,
    verify_license_online,
    log_usage_online,
)


# ================= FIX SSL =================
ssl._create_default_https_context = ssl._create_unverified_context

CURRENT_VERSION = "1.15"
UPDATE_MANIFEST_URL = "https://raw.githubusercontent.com/ibrahim-saiied/Bedayti---Public-Prosecution-Automation/clean-main/version.json"


# ================= GUI =================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Bedayti - Public Prosecution Automation v1.15")
        self.geometry("820x370")
        self.resizable(False, False)
        self.configure(bg="#eef2f7")

        self.excel_path = tk.StringVar()
        self.pdf_path = tk.StringVar()
        self.license_key = tk.StringVar(value=(read_saved_license_key() or "TEST-1234"))
        self.otp_code = tk.StringVar()
        self.captcha_code = tk.StringVar()
        self.attach_existing_browser = tk.BooleanVar(value=False)
        self.debugger_address = tk.StringVar(value="127.0.0.1:9222")
        self.script_dir = get_runtime_dir()
        self.status_var = tk.StringVar(value="جاهز")
        self.service_options = [
            "التقديم علي احكام جنائي",
            "التقديم علي محاضر شرطة",
            "الاستعلام عن حالة الطلب",
        ]
        self.selected_service = tk.StringVar(value=self.service_options[0])
        self.automation_services = {
            "التقديم علي احكام جنائي",
            "التقديم علي محاضر شرطة",
        }
        self.inquiry_rows = []
        self.ocr_warned = False
        self.ocr_ready = False
        self.setup_ocr_engine()

        self.form_profiles = {
            "التقديم علي احكام جنائي": {
                "request_ready": (By.NAME, "P23_CAUSE_NUMBER"),
                "case_number": (By.NAME, "P23_CAUSE_NUMBER"),
                "case_year": (By.NAME, "P23_CAUSE_YEAR"),
                "table": (By.ID, "P23_TABLE"),
                "gov": (By.ID, "P23_GOV"),
                "police_department": (By.ID, "P23_POLICE_DEPARTMENT"),
                "send_to": (By.ID, "P23_SEND_TO"),
                "agent_description": (By.ID, "P23_AGENT_DESCRIPTION"),
                "agent_number": (By.ID, "P23_AGENT_NUMBER"),
                "card_number": (By.ID, "P23_CARD_NUMBER"),
                "entry_type": (By.ID, "P23_ENTRY_TYPE"),
                "entity": (By.ID, "P23_ENTITY"),
                "add_client_btn": (By.ID, "B1"),
                "client_petitioner_desc": (By.ID, "P26_PETITIONER_DESC"),
                "client_identity_type": (By.ID, "P26_IDENTITY_TYPE"),
                "client_national_id": (By.ID, "P26_NATIONAL_ID"),
                "client_first_name": (By.ID, "P26_FIRST_NAME"),
                "client_second_name": (By.ID, "P26_SECOND_NAME"),
                "client_third_name": (By.ID, "P26_THIRD_NAME"),
                "client_fourth_name": (By.ID, "P26_FOURTH_NAME"),
                "client_address": (By.ID, "P26_ADDRESS"),
                "client_email": (By.ID, "P26_EMAIL"),
                "dialog_save_btn": (By.ID, "B3"),
                "doc_dialog_frame_css": "iframe[src*='add-attatchment']",
                "attachment_type": (By.ID, "P21_ATTATCHMENT_TYPE"),
                "attachment_input": (By.ID, "P21_ATTATCHMENT_input"),
                "delivery_receipt": (By.ID, "P23_RECEIPT"),
                "delivery_gov": (By.ID, "P23_DELIVERY_GOV"),
                "delivery_phone": (By.ID, "P23_CONTACT_PHONE_NUMBER"),
                "delivery_address": (By.ID, "P23_DELIVERY_ADD"),
                "terms_checkbox": (By.ID, "P23_TERMS_CONDITIONS_LABEL"),
                "captcha_field": (By.ID, "P23_CAPTCHA_CODE"),
                "captcha_fallback_css": "input[name*='captcha'], input[id*='captcha']",
                "submit_button": (By.ID, "cid-submit"),
                "series_refresh": (By.CSS_SELECTOR, ".fas.fa-sync-alt.cc-refresh"),
                "series_display": (By.ID, "P40_SERIES_DISPLAY"),
                "series_inprogress_text": "جارى إنشاء الطلب.",
            },
            "التقديم علي محاضر شرطة": {
                # TODO: عدّل القيم التالية حسب IDs الفعلية لخدمة محاضر الشرطة
                "request_ready": (By.NAME, "P23_CAUSE_NUMBER"),
                "case_number": (By.NAME, "P23_CAUSE_NUMBER"),
                "case_year": (By.NAME, "P23_CAUSE_YEAR"),
                "table": (By.ID, "P23_TABLE"),
                "gov": (By.ID, "P23_GOV"),
                "police_department": (By.ID, "P23_POLICE_DEPARTMENT"),
                "send_to": (By.ID, "P23_SEND_TO"),
                "agent_description": (By.ID, "P23_AGENT_DESCRIPTION"),
                "agent_number": (By.ID, "P23_AGENT_NUMBER"),
                "card_number": (By.ID, "P23_CARD_NUMBER"),
                "entry_type": (By.ID, "P23_ENTRY_TYPE"),
                "entity": (By.ID, "P23_ENTITY"),
                "add_client_btn": (By.ID, "B1"),
                "client_petitioner_desc": (By.ID, "P26_PETITIONER_DESC"),
                "client_identity_type": (By.ID, "P26_IDENTITY_TYPE"),
                "client_national_id": (By.ID, "P26_NATIONAL_ID"),
                "client_first_name": (By.ID, "P26_FIRST_NAME"),
                "client_second_name": (By.ID, "P26_SECOND_NAME"),
                "client_third_name": (By.ID, "P26_THIRD_NAME"),
                "client_fourth_name": (By.ID, "P26_FOURTH_NAME"),
                "client_address": (By.ID, "P26_ADDRESS"),
                "client_email": (By.ID, "P26_EMAIL"),
                "dialog_save_btn": (By.ID, "B3"),
                "doc_dialog_frame_css": "iframe[src*='add-attatchment']",
                "attachment_type": (By.ID, "P21_ATTATCHMENT_TYPE"),
                "attachment_input": (By.ID, "P21_ATTATCHMENT_input"),
                "delivery_receipt": (By.ID, "P23_RECEIPT"),
                "delivery_gov": (By.ID, "P23_DELIVERY_GOV"),
                "delivery_phone": (By.ID, "P23_CONTACT_PHONE_NUMBER"),
                "delivery_address": (By.ID, "P23_DELIVERY_ADD"),
                "terms_checkbox": (By.ID, "P23_TERMS_CONDITIONS_LABEL"),
                "captcha_field": (By.ID, "P23_CAPTCHA_CODE"),
                "captcha_fallback_css": "input[name*='captcha'], input[id*='captcha']",
                "submit_button": (By.ID, "cid-submit"),
                "series_refresh": (By.CSS_SELECTOR, ".fas.fa-sync-alt.cc-refresh"),
                "series_display": (By.ID, "P40_SERIES_DISPLAY"),
                "series_inprogress_text": "جارى إنشاء الطلب.",
            },
            "الاستعلام عن حالة الطلب": {
                "request_ready": (By.NAME, "P29_PETITION_SERIES"),
                "inquiry_series_field": (By.NAME, "P29_PETITION_SERIES"),
                "captcha_field": (By.NAME, "P29_CODE"),
                "captcha_fallback_css": "input[name='P29_CODE'], input[id*='P29_CODE'], input[name*='CODE']",
                "submit_button": (By.ID, "B3176408409268685919"),
            },
        }

        # Header
        self.header_frame = tk.Frame(self, bg="#0f172a", height=78)
        self.header_frame.pack(fill=tk.X)
        self.header_frame.pack_propagate(False)

        self.header_left = tk.Frame(self.header_frame, bg="#0f172a")
        self.header_left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.header_right = tk.Frame(self.header_frame, bg="#0f172a")
        self.header_right.pack(side=tk.RIGHT, padx=14, pady=8)

        tk.Label(
            self.header_left,
            text="Bedayti - Public Prosecution Automation System v1.15",
            font=("Segoe UI", 16, "bold"),
            bg="#0f172a",
            fg="white"
        ).pack(anchor="w", padx=16, pady=(10, 0))
        tk.Label(
            self.header_left,
            text="تسجيل البيانات تلقائيا علي موقع النيابة العامة",
            font=("Segoe UI", 10),
            bg="#0f172a",
            fg="#cbd5e1"
        ).pack(anchor="w", padx=16, pady=(0, 8))

        tk.Label(
            self.header_right,
            text="✦ Ibrahim Saiied ✦",
            bg="#0f172a",
            fg="#fbbf24",
            font=("Segoe UI", 11, "bold")
        ).pack(anchor="e")
        tk.Label(
            self.header_right,
            text="☎ 01144366104       ",
            bg="#0f172a",
            fg="#e2e8f0",
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="e")

        # Main card
        self.main_card = tk.Frame(self, bg="white", bd=1, relief="solid", padx=14, pady=12)
        self.main_card.pack(fill=tk.BOTH, expand=True, padx=14, pady=12)

        self.files_frame = tk.LabelFrame(
            self.main_card, text="الملفات", bg="white", fg="#0f172a", font=("Segoe UI", 10, "bold"), padx=10, pady=8
        )
        self.files_frame.pack(fill=tk.X, pady=(0, 8))

        service_row = tk.Frame(self.files_frame, bg="white")
        service_row.pack(fill=tk.X, pady=4)
        tk.Label(service_row, text="نوع الخدمة", bg="white").pack(side=tk.LEFT, padx=(0, 8))
        tk.OptionMenu(service_row, self.selected_service, *self.service_options).pack(side=tk.LEFT)
        tk.Entry(service_row, textvariable=self.license_key, width=30, relief="solid", bd=1).pack(side=tk.RIGHT, padx=(0, 8))
        tk.Label(service_row, text="License", bg="white").pack(side=tk.RIGHT, padx=(0, 8))

        row1 = tk.Frame(self.files_frame, bg="white")
        row1.pack(fill=tk.X, pady=4)
        tk.Label(row1, text="ملف Excel", bg="white").pack(side=tk.LEFT, padx=(0, 8))
        tk.Entry(row1, textvariable=self.excel_path, width=75, relief="solid", bd=1).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(row1, text="اختيار Excel", command=self.select_excel, bg="#e2e8f0").pack(side=tk.LEFT)

        row2 = tk.Frame(self.files_frame, bg="white")
        row2.pack(fill=tk.X, pady=4)
        tk.Label(row2, text="ملف PDF (التوكيل)", bg="white").pack(side=tk.LEFT, padx=(0, 8))
        tk.Entry(row2, textvariable=self.pdf_path, width=68, relief="solid", bd=1).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(row2, text="اختيار PDF", command=self.select_pdf, bg="#e2e8f0").pack(side=tk.LEFT)

        browser_row = tk.Frame(self.files_frame, bg="white")
        browser_row.pack(fill=tk.X, pady=4)
        tk.Checkbutton(
            browser_row,
            text="استخدم نافذة Chrome المفتوحة بعد تسجيل الدخول",
            variable=self.attach_existing_browser,
            bg="white",
            activebackground="white",
        ).pack(side=tk.LEFT, padx=(0, 8))
        tk.Label(browser_row, text="Debugger", bg="white").pack(side=tk.LEFT, padx=(14, 6))
        tk.Entry(browser_row, textvariable=self.debugger_address, width=22, relief="solid", bd=1).pack(side=tk.LEFT)

        browser_hint_row = tk.Frame(self.files_frame, bg="white")
        browser_hint_row.pack(fill=tk.X, pady=(0, 2))
        tk.Label(
            browser_hint_row,
            text="مهم: يجب أن تكون نافذة Chrome المفتوحة شغالة بـ remote debugging مثل 127.0.0.1:9222",
            bg="white",
            fg="#64748b",
            anchor="w",
            justify="left",
        ).pack(fill=tk.X)

        self.actions_frame = tk.Frame(self.main_card, bg="white")
        self.actions_frame.pack(fill=tk.X, pady=(2, 10))
        self.start_btn = tk.Button(
            self.actions_frame,
            text="يلا بينا",
            command=self.start,
            bg="#16a34a",
            fg="white",
            activebackground="#15803d",
            activeforeground="white",
            font=("Segoe UI", 10, "bold"),
            relief="raised",
            bd=2,
            padx=19,
            pady=4
        )
        self.start_btn.pack(anchor="center")

        self.reload_btn = tk.Button(
            self.actions_frame,
            text="إعادة تحميل الطلب الحالي",
            command=self.reload_current,
            state=tk.DISABLED,
            bg="#f59e0b",
            fg="white",
            activebackground="#d97706",
            activeforeground="white",
            padx=12
        )

        self.otp_frame = tk.LabelFrame(
            self.main_card, text="OTP", bg="white", fg="#0f172a", font=("Segoe UI", 10, "bold"), padx=10, pady=8
        )
        tk.Label(self.otp_frame, text="OTP", bg="white").pack(side=tk.LEFT)
        tk.Entry(self.otp_frame, textvariable=self.otp_code, width=24, relief="solid", bd=1).pack(side=tk.LEFT, padx=6)
        tk.Button(self.otp_frame, text="تأكيد", command=self.submit_otp, bg="#10b981", fg="white").pack(side=tk.LEFT)

        self.step_frame = tk.LabelFrame(
            self.main_card, text="الكابتشا", bg="white", fg="#0f172a", font=("Segoe UI", 10, "bold"), padx=10, pady=8
        )
        tk.Label(self.step_frame, text="ادخل الكابتشا:", bg="white").pack(side=tk.LEFT)
        tk.Entry(self.step_frame, textvariable=self.captcha_code, width=22, relief="solid", bd=1).pack(side=tk.LEFT, padx=6)
        self.next_btn = tk.Button(
            self.step_frame,
            text="الخطوة التالية",
            command=self.next_step,
            state=tk.DISABLED,
            bg="#2563eb",
            fg="white",
            activebackground="#1d4ed8",
            activeforeground="white"
        )
        self.next_btn.pack(side=tk.LEFT)

        self.status_frame = tk.Frame(self.main_card, bg="white")
        self.status_frame.pack(fill=tk.X, pady=(0, 6))
        self.status_label = tk.Label(
            self.status_frame,
            textvariable=self.status_var,
            bg="white",
            fg="#334155",
            anchor="w",
            justify="left"
        )
        self.status_label.pack(fill=tk.X)
        self.update_progress_var = tk.DoubleVar(value=0.0)
        self.update_progress_bar = ttk.Progressbar(
            self.status_frame,
            orient="horizontal",
            mode="determinate",
            variable=self.update_progress_var,
            maximum=100
        )
        self.update_progress_bar.pack(fill=tk.X, pady=(4, 0))
        self.update_progress_bar.pack_forget()

        self.driver = None
        self.wait = None
        self.fixed = {}
        self.cases = None
        self.current_index = 0
        self.case_tabs = []
        self.batch_case_indices = []
        self.current_batch_start = 0
        self.max_open_tabs = 50
        self.request_url = ""
        self.session_submitted_requests = 0
        self.session_counted_case_indices = set()
        self.state = "idle"
        self.set_default_file_paths()
        self.bind_all("<Return>", self.on_enter_pressed)
        self.after(900, self.check_for_updates_silent)

    # ================= HELPERS =================
    def err(self, msg, raise_exc=True):
        messagebox.showerror("خطأ", msg)
        if raise_exc:
            raise Exception(msg)

    def on_enter_pressed(self, event=None):
        try:
            if self.otp_frame.winfo_ismapped():
                self.submit_otp()
                return "break"
            if self.step_frame.winfo_ismapped() and str(self.next_btn.cget("state")) != "disabled":
                self.next_step()
                return "break"
        except Exception:
            pass
        return None

    def ui_sleep(self, seconds, tick=0.05):
        end_time = time.monotonic() + max(0.0, float(seconds))
        while time.monotonic() < end_time:
            try:
                self.update_idletasks()
                self.update()
            except Exception:
                pass
            time.sleep(tick)

    def parse_version(self, version_text):
        parts = re.findall(r"\d+", str(version_text or ""))
        if not parts:
            return (0,)
        return tuple(int(p) for p in parts)

    def is_newer_version(self, latest, current):
        return self.parse_version(latest) > self.parse_version(current)

    def fetch_update_manifest(self, timeout=10):
        req = urllib.request.Request(UPDATE_MANIFEST_URL, method="GET")
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8", errors="ignore")
        return json.loads(raw)

    def _hide_update_progress(self):
        try:
            if self.update_progress_bar.winfo_ismapped():
                self.update_progress_bar.pack_forget()
            self.update_progress_var.set(0.0)
        except Exception:
            pass

    def check_for_updates_silent(self):
        if getattr(self, "_update_check_started", False):
            return
        self._update_check_started = True
        t = threading.Thread(target=self._update_worker, daemon=True)
        t.start()

    def _notify_update_link(self, latest, update_url):
        message = (
            f"يوجد إصدار جديد: {latest}\n\n"
            "سيتم فتح رابط التحميل المباشر الآن.\n"
            "هل تريد فتح صفحة التحميل؟"
        )
        open_now = messagebox.askyesno("تحديث جديد", message)
        if open_now:
            try:
                webbrowser.open(update_url, new=2)
                self.status_var.set(f"تم فتح رابط التحديث: v{latest}")
            except Exception as e:
                self.status_var.set(f"تعذر فتح رابط التحديث: {e}")
                messagebox.showerror("تحديث", f"تعذر فتح الرابط تلقائيًا:\n{update_url}")
        else:
            self.status_var.set(f"يوجد تحديث v{latest} متاح للتحميل.")
            messagebox.showinfo("رابط التحديث", f"رابط التحميل المباشر:\n{update_url}")

    def _update_worker(self):
        try:
            manifest = self.fetch_update_manifest(timeout=8)
            latest = str(manifest.get("version", "")).strip()
            if not latest or not self.is_newer_version(latest, CURRENT_VERSION):
                return

            update_url = str(manifest.get("url", "")).strip()
            if not update_url:
                self.after(0, lambda: messagebox.showerror("تحديث", "رابط التحديث غير متاح."))
                return

            self.after(0, lambda l=latest, u=update_url: self._notify_update_link(l, u))
        except Exception as e:
            self.after(0, self._hide_update_progress)
            self.after(0, lambda msg=str(e): self.status_var.set(f"فشل فحص التحديث: {msg}"))
            return

    def hide_prestart_widgets(self):
        for w in (self.files_frame,):
            try:
                w.pack_forget()
            except Exception:
                pass
        try:
            self.start_btn.config(state=tk.DISABLED)
            self.start_btn.pack_forget()
        except Exception:
            pass

    def show_reload_button(self):
        try:
            if not self.reload_btn.winfo_ismapped():
                self.reload_btn.pack(side=tk.LEFT, padx=(8, 0))
            self.reload_btn.config(state=tk.NORMAL)
        except Exception:
            pass

    def hide_runtime_widgets_on_finish(self):
        try:
            self.step_frame.pack_forget()
        except Exception:
            pass
        try:
            self.otp_frame.pack_forget()
        except Exception:
            pass
        try:
            self.reload_btn.config(state=tk.DISABLED)
            self.reload_btn.pack_forget()
        except Exception:
            pass

    def build_driver(self, options, debugger_address=""):
        if debugger_address:
            options.add_experimental_option("debuggerAddress", debugger_address)

        local_driver = self.script_dir / "chromedriver.exe"
        if local_driver.is_file():
            service = Service(str(local_driver))
            return webdriver.Chrome(service=service, options=options)

        path_driver = shutil.which("chromedriver")
        if path_driver:
            try:
                service = Service(path_driver)
                return webdriver.Chrome(service=service, options=options)
            except Exception:
                pass

        try:
            return webdriver.Chrome(options=options)
        except Exception:
            pass

        os.environ.setdefault("WDM_SSL_VERIFY", "0")
        service = Service(ChromeDriverManager().install())
        return webdriver.Chrome(service=service, options=options)

    def normalize_debugger_address(self):
        address = str(self.debugger_address.get() or "").strip()
        address = address.replace("http://", "").replace("https://", "").strip().strip("/")
        if not address:
            self.err("ادخل عنوان Debugger مثل 127.0.0.1:9222")
        return address

    def check_debugger_endpoint(self, address, timeout=2):
        try:
            url = f"http://{address}/json/version"
            req = urllib.request.Request(url, method="GET")
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                raw = resp.read().decode("utf-8", errors="ignore")
            data = json.loads(raw)
            return bool(data.get("Browser"))
        except Exception:
            return False

    def has_locator(self, locator):
        try:
            return bool(self.driver.find_elements(*locator))
        except Exception:
            return False

    def is_service_selection_page(self):
        try:
            labels = self.get_service_labels()
            return bool(self.driver.execute_script(
                """
                const labels = arguments[0] || [];
                const texts = Array.from(document.querySelectorAll('body, a, button, [role="button"], .a-CardView, .a-CardView-item, .t-Card'))
                    .map(el => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim())
                    .filter(Boolean);
                return labels.some(label => texts.some(txt => txt.includes(label)));
                """,
                labels,
            ))
        except Exception:
            return False

    def get_selected_service(self):
        return (self.selected_service.get() or "").strip()

    def is_automation_service(self):
        return self.get_selected_service() in self.automation_services

    def get_form_profile(self):
        selected = self.get_selected_service()
        return self.form_profiles.get(selected, self.form_profiles["التقديم علي احكام جنائي"])

    def get_profile_value(self, key):
        profile = self.get_form_profile()
        if key not in profile:
            self.err(f"المتغير [{key}] غير موجود في إعدادات الخدمة: {self.get_selected_service()}")
        return profile[key]

    def get_profile_locator(self, key):
        locator = self.get_profile_value(key)
        if not isinstance(locator, tuple) or len(locator) != 2:
            self.err(f"المتغير [{key}] يجب أن يكون locator بشكل (By.*, 'value').")
        return locator

    def is_inquiry_service(self):
        return self.get_selected_service() == "الاستعلام عن حالة الطلب"

    def read_excel_sheet(self, sheet_name):
        try:
            return pd.read_excel(self.excel_path.get(), sheet_name=sheet_name)
        except ValueError:
            self.err(f"لم يتم العثور على الشيت [{sheet_name}] داخل ملف Excel.")
        except Exception as e:
            self.err(f"تعذر قراءة الشيت [{sheet_name}] من ملف Excel: {e}")

    def ensure_sheet_columns(self, df, sheet_name, required_columns):
        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            self.err(
                f"الشيت [{sheet_name}] ناقص الأعمدة التالية:\n"
                f"{', '.join(missing)}"
            )

    def load_input_data(self, include_cases=True):
        if not self.excel_path.get():
            self.err("اختار ملف Excel")
        if not os.path.isfile(self.excel_path.get()):
            self.err("مسار ملف Excel غير موجود")

        df_fixed = self.read_excel_sheet("Fixed_Data")
        self.ensure_sheet_columns(df_fixed, "Fixed_Data", ["الحقل", "البيانات"])
        self.fixed = dict(zip(df_fixed["الحقل"], df_fixed["البيانات"]))

        if not include_cases:
            self.cases = pd.DataFrame()
            return

        if self.is_automation_service():
            if not self.pdf_path.get():
                self.err("اختار ملف PDF")
            if not os.path.isfile(self.pdf_path.get()):
                self.err("مسار ملف PDF غير موجود")

        self.cases = self.read_excel_sheet("Cases_Data")
        if self.cases.empty:
            self.err("Cases_Data فاضي")
        if self.is_automation_service():
            self.ensure_sheet_columns(
                self.cases,
                "Cases_Data",
                ["رقم_القضية", "سنة_القضية", "الجدول", "المحافظة", "قسم_الشرطة", "الي"],
            )
        self.ensure_request_number_column()

    def ensure_request_number_column(self):
        if "رقم_الطلب" not in self.cases.columns:
            self.cases["رقم_الطلب"] = ""
            return

        col = self.cases["رقم_الطلب"]
        self.cases["رقم_الطلب"] = col.astype("object")
        self.cases.loc[col.isna(), "رقم_الطلب"] = ""

    def ensure_inquiry_status_column(self):
        if "حالة_الطلب" not in self.cases.columns:
            self.cases["حالة_الطلب"] = ""
            return

        col = self.cases["حالة_الطلب"]
        self.cases["حالة_الطلب"] = col.astype("object")
        self.cases.loc[col.isna(), "حالة_الطلب"] = ""

    def get_fixed(self, key):
        if key not in self.fixed or pd.isna(self.fixed[key]):
            self.err(f"الحقل [{key}] ناقص في Fixed_Data")
        return str(self.fixed[key]).strip()

    def get_case(self, row, col):
        if col not in row or pd.isna(row[col]):
            self.err(f"الحقل [{col}] ناقص في Cases_Data\nصف {row.name+2}")
        value = str(row[col]).strip()
        if col in ("رقم_القضية", "سنة_القضية") and re.fullmatch(r"\d+\.0", value):
            value = value[:-2]
        if col == "المحافظة":
            value = value.replace("_", " ")
        return value

    def get_captcha_input_element(self, primary_timeout=2, fallback_timeout=4):
        try:
            return WebDriverWait(self.driver, primary_timeout).until(
                EC.presence_of_element_located(self.get_profile_locator("captcha_field"))
            )
        except Exception:
            return WebDriverWait(self.driver, fallback_timeout).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, self.get_profile_value("captcha_fallback_css")))
            )

    def clear_and_type(self, locator_by, locator_value, text):
        e = self.wait.until(EC.presence_of_element_located((locator_by, locator_value)))
        try:
            e.clear()
        except Exception:
            pass
        e.send_keys(text)

    def select_option_fuzzy(self, locator_by, locator_value, text):
        select = Select(self.wait.until(EC.presence_of_element_located((locator_by, locator_value))))
        all_options = [option.text.strip() for option in select.options]

        if text in all_options:
            select.select_by_visible_text(text)
            return

        matches = get_close_matches(text, all_options, n=1, cutoff=0.6)
        if matches:
            select.select_by_visible_text(matches[0])
            return

        self.err(f"لم يُعثر على '{text}' في الخيارات المتاحة:\n{', '.join(all_options)}")

    def wait_dropdown_loaded(self, locator_by, locator_value, expected_text=None, timeout=12):
        def _is_loaded(_):
            try:
                elem = self.driver.find_element(locator_by, locator_value)
                opts = [o.text.strip() for o in Select(elem).options if o.text.strip()]
                if not opts:
                    return False

                normalized = [o for o in opts if o not in ("اختر", "اختر...", "-- اختر --")]
                if not normalized:
                    return False

                if expected_text:
                    if expected_text in normalized:
                        return True
                    return bool(get_close_matches(expected_text, normalized, n=1, cutoff=0.6))

                return len(normalized) >= 2
            except Exception:
                return False

        WebDriverWait(self.driver, timeout).until(_is_loaded)

    def switch_to_dialog_frame(self, frame_css, field_locator, timeout=12):
        d = self.driver
        if isinstance(field_locator, tuple) and len(field_locator) == 2:
            target_locator = field_locator
        else:
            target_locator = (By.ID, str(field_locator))

        try:
            d.switch_to.default_content()
        except Exception:
            pass

        try:
            WebDriverWait(d, timeout).until(
                EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR, frame_css))
            )
            WebDriverWait(d, timeout).until(EC.presence_of_element_located(target_locator))
            return
        except Exception:
            d.switch_to.default_content()

        end_time = time.time() + timeout
        while time.time() < end_time:
            iframes = d.find_elements(By.TAG_NAME, "iframe")
            for frm in iframes:
                try:
                    d.switch_to.default_content()
                    d.switch_to.frame(frm)
                    if d.find_elements(*target_locator):
                        return
                except Exception:
                    pass
            self.ui_sleep(0.2)
        d.switch_to.default_content()
        raise TimeoutException(f"لم يتم العثور على iframe يحتوي locator: {target_locator}")

    def find_latest_file(self, patterns):
        search_dirs = [self.script_dir]
        cwd = Path.cwd().resolve()
        if cwd != self.script_dir:
            search_dirs.append(cwd)

        for base_dir in search_dirs:
            candidates = []
            for pattern in patterns:
                candidates.extend(base_dir.glob(pattern))

            files = [f for f in candidates if f.is_file()]
            if files:
                latest = max(files, key=lambda f: f.stat().st_mtime)
                return str(latest.resolve())

        return ""

    def set_default_file_paths(self):
        excel_default = self.find_latest_file(["*.xlsx", "*.xls"])
        pdf_default = self.find_latest_file(["*.pdf"])

        if excel_default:
            self.excel_path.set(excel_default)
            self.load_license_key_from_excel(excel_default)
        if pdf_default:
            self.pdf_path.set(pdf_default)

    def load_license_key_from_excel(self, excel_path):
        try:
            if not excel_path or not os.path.isfile(excel_path):
                return
            df_fixed = pd.read_excel(excel_path, sheet_name="Fixed_Data")
            if "الحقل" not in df_fixed.columns or "البيانات" not in df_fixed.columns:
                return
            fixed_map = dict(zip(df_fixed["الحقل"], df_fixed["البيانات"]))
            val = fixed_map.get("license_key", "")
            if pd.isna(val):
                return
            val = str(val).strip()
            if val:
                self.license_key.set(val)
        except Exception:
            pass

    def get_service_labels(self):
        selected = self.get_selected_service()
        labels_map = {
            "التقديم علي احكام جنائي": [
                "التقديم علي احكام جنائي",
                "التقديم على أحكام جنائي",
                "أحكام جنائي",
            ],
            "التقديم علي محاضر شرطة": [
                "التقديم علي محاضر شرطة",
                "التقديم على محاضر شرطة",
                "محاضر شرطة",
            ],
            "الاستعلام عن حالة الطلب": [
                "الاستعلام عن حالة الطلب",
                "حالة الطلب",
            ],
        }
        return labels_map.get(selected, [selected])

    def run_automation_after_login(self):
        w = self.wait
        w.until(EC.presence_of_element_located(self.get_profile_locator("request_ready")))
        self.status_var.set("تم فتح صفحة الطلبات")

        try:
            self.otp_frame.pack_forget()
        except Exception:
            pass

        self.show_reload_button()
        self.current_index = 0
        self.case_tabs = []
        self.batch_case_indices = []
        self.current_batch_start = 0
        self.request_url = self.driver.current_url
        self.session_submitted_requests = 0
        self.session_counted_case_indices = set()
        self.prepare_all_cases_tabs()

    def run_inquiry_after_login(self):
        w = self.wait
        w.until(EC.presence_of_element_located(self.get_profile_locator("request_ready")))
        self.status_var.set("تم فتح صفحة الاستعلام")

        try:
            self.otp_frame.pack_forget()
        except Exception:
            pass

        self.show_reload_button()
        self.current_index = 0
        self.case_tabs = []
        self.inquiry_rows = []
        self.request_url = self.driver.current_url
        self.prepare_all_inquiry_tabs()

    def open_selected_service_card(self):
        d = self.driver
        selected = self.get_selected_service()

        data_id_map = {
            "التقديم علي احكام جنائي": "4",
            "التقديم علي محاضر شرطة": "2",
            "الاستعلام عن حالة الطلب": "5",
        }
        target_data_id = data_id_map.get(selected)
        if target_data_id is not None:
            end_time = time.time() + 12
            while time.time() < end_time:
                clicked = d.execute_script(
                    """
                    const dataId = arguments[0];
                    const root = document.querySelector(`[data-id="${dataId}"]`);
                    if (!root) return false;

                    const clickEl = root.matches('a,button,[role="button"]')
                        ? root
                        : (root.querySelector('a,button,[role="button"]') || root);
                    clickEl.click();
                    return true;
                    """,
                    target_data_id
                )
                if clicked:
                    return
                self.ui_sleep(0.3)

        labels = self.get_service_labels()
        end_time = time.time() + 12

        while time.time() < end_time:
            matched = d.execute_script(
                """
                const labels = arguments[0] || [];
                const cardSelectors = ['.a-CardView', '.a-CardView-item', '.t-Card'];
                const cards = cardSelectors.flatMap(s => Array.from(document.querySelectorAll(s)));

                for (const label of labels) {
                    for (const card of cards) {
                        const txt = (card.innerText || card.textContent || '').replace(/\\s+/g, ' ').trim();
                        if (!txt.includes(label)) continue;
                        const clickEl = card.querySelector('a,button,[role="button"]') || card.closest('a,button');
                        if (clickEl) {
                            clickEl.click();
                            return label;
                        }
                    }

                    const direct = Array.from(document.querySelectorAll('a,button,[role="button"]')).find(el => {
                        const txt = (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
                        return txt.includes(label);
                    });
                    if (direct) {
                        direct.click();
                        return label;
                    }
                }
                return '';
                """,
                labels
            )
            if matched:
                return
            self.ui_sleep(0.3)

        # ===== FIX 1: صُحّح الـ CSS selector — كان "a-CardView-fullLinka-CardView-fullLink" بدون نقطة =====
        if selected == "التقديم علي احكام جنائي":
            self.wait.until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, ".a-CardView-fullLink")
                )
            ).click()
            return

        self.err(f"لم يتم العثور على بطاقة الخدمة المختارة: {selected}")

    # ================= FILE PICKERS =================
    def select_excel(self):
        selected = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
        )
        if not selected:
            return
        self.excel_path.set(selected)
        self.load_license_key_from_excel(selected)

    def select_pdf(self):
        selected = filedialog.askopenfilename(
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )
        if selected:
            self.pdf_path.set(selected)

    # ================= START =================
    def start(self):
        try:
            key = (self.license_key.get() or "").strip()
            if not key:
                self.err("ادخل License أولًا.")
            self.status_var.set("جاري التحقق من الترخيص...")
            ok, reason = verify_license_online(key, timeout=7)
            if not ok:
                reason_map = {
                    "network_error": "تعذر الوصول إلى خادم التحقق. تحقق من الإنترنت ثم أعد المحاولة.",
                    "verify_error": "حدث خطأ مؤقت أثناء التحقق من License. أعد المحاولة بعد لحظات.",
                }
                self.err(f"فشل التحقق من License: {reason_map.get(reason, reason)}")
            save_license_key(key)
            self.session_submitted_requests = 0
            self.session_counted_case_indices = set()
            self.status_var.set("جاري قراءة بيانات Excel...")
            self.load_input_data(include_cases=(self.is_automation_service() or self.is_inquiry_service()))

            options = Options()
            if self.attach_existing_browser.get():
                debugger_address = self.normalize_debugger_address()
                if not self.check_debugger_endpoint(debugger_address):
                    self.err(
                        "تعذر الاتصال بواجهة Chrome Debugger.\n"
                        "تأكد أن نافذة Chrome الحالية تم تشغيلها بـ remote debugging على نفس العنوان.\n"
                        "مثال: chrome.exe --remote-debugging-port=9222 --user-data-dir=%TEMP%\\ppo-debug-profile"
                    )
                self.status_var.set("جاري الاتصال بمتصفح Chrome المفتوح...")
                self.driver = self.build_driver(options, debugger_address=debugger_address)
            else:
                options.add_argument("--start-maximized")
                options.add_experimental_option("detach", True)
                self.status_var.set("جاري فتح المتصفح...")
                self.driver = self.build_driver(options)

            self.wait = WebDriverWait(self.driver, 30)

            if self.attach_existing_browser.get():
                self.hide_prestart_widgets()
                self.status_var.set("تم الاتصال بمتصفح Chrome المفتوح. جارٍ المتابعة بعد تسجيل الدخول...")
                self.after_login()
                return

            self.driver.get("https://www.ppo.gov.eg/ppo/r/ppoportal/ppoportal/login-page")

            self.wait.until(EC.presence_of_element_located((By.NAME, "P9999_USERNAME"))).send_keys(self.get_fixed("اسم_المستخدم"))
            self.wait.until(EC.presence_of_element_located((By.NAME, "P9999_PASSWORD"))).send_keys(self.get_fixed("الرقم_السري"))
            self.wait.until(EC.element_to_be_clickable((By.ID, "GENERATE_OTP"))).click()

            self.hide_prestart_widgets()
            self.otp_frame.pack(fill=tk.X, pady=8)
            self.status_var.set("انتظر إدخال OTP")

        except Exception as e:
            self.err(f"فشل البدء: {e}", raise_exc=False)

    # ================= OTP =================
    def submit_otp(self):
        try:
            self.wait.until(EC.presence_of_element_located((By.ID, "P9999_OTP_VER"))).send_keys(self.otp_code.get())
            self.wait.until(EC.element_to_be_clickable((By.ID, "LOGIN"))).click()
            self.wait.until(EC.presence_of_element_located((By.ID, "navbarDropdownMenuLink")))
            self.after_login()

        except Exception as e:
            self.err(f"فشل OTP: {e}", raise_exc=False)

    # ================= AFTER LOGIN =================
    def after_login(self, suppress_error=False):
        w = self.wait
        try:
            if not self.has_locator(self.get_profile_locator("request_ready")):
                navigated_from_menu = False
                try:
                    w.until(EC.element_to_be_clickable((By.ID, "navbarDropdownMenuLink"))).click()
                    w.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".dropdown-menu .dropdown-item:nth-child(2)"))).click()
                    navigated_from_menu = True
                except Exception:
                    pass

                if not self.has_locator(self.get_profile_locator("request_ready")):
                    if navigated_from_menu or self.is_service_selection_page():
                        self.open_selected_service_card()
                    else:
                        self.err(
                            "تعذر تحديد نقطة البداية بعد تسجيل الدخول.\n"
                            "افتح صفحة الخدمات أو صفحة الخدمة المطلوبة ثم أعد المحاولة."
                        )

            if self.is_automation_service():
                self.run_automation_after_login()
                return True
            if self.is_inquiry_service():
                self.run_inquiry_after_login()
                return True

            self.err(f"نوع خدمة غير مدعوم: {self.get_selected_service()}")
            return False

        except Exception as e:
            if not suppress_error:
                self.err(f"فشل بعد تسجيل الدخول: {e}", raise_exc=False)
            return False

    # ================= MULTI-TAB CASE PREPARATION =================
    def prepare_case_in_current_tab(self, row, idx, total):
        d, w = self.driver, self.wait
        self.status_var.set(f"تجهيز الطلب {idx+1}/{total} | رقم الطلب: {row['رقم_الطلب']}")

        self.clear_and_type(*self.get_profile_locator("case_number"), self.get_case(row, "رقم_القضية"))
        self.clear_and_type(*self.get_profile_locator("case_year"), self.get_case(row, "سنة_القضية"))
        self.select_option_fuzzy(*self.get_profile_locator("table"), self.get_case(row, "الجدول"))
        self.select_option_fuzzy(*self.get_profile_locator("gov"), self.get_case(row, "المحافظة"))
        police_department = self.get_case(row, "قسم_الشرطة")
        self.wait_dropdown_loaded(*self.get_profile_locator("police_department"), expected_text=police_department, timeout=12)
        self.select_option_fuzzy(*self.get_profile_locator("police_department"), police_department)
        self.select_option_fuzzy(*self.get_profile_locator("send_to"), self.get_case(row, "الي"))

        self.select_option_fuzzy(*self.get_profile_locator("agent_description"), self.get_fixed("توصيف_الوكيل"))
        self.clear_and_type(*self.get_profile_locator("agent_number"), self.get_fixed("رقم_التوكيل"))
        self.clear_and_type(*self.get_profile_locator("card_number"), self.get_fixed("رقم_الكارنية"))
        self.select_option_fuzzy(*self.get_profile_locator("entry_type"), self.get_fixed("نوع_القيد"))
        self.clear_and_type(*self.get_profile_locator("entity"), self.get_fixed("جهة_إصدار_التوكيل"))

        # ===== FIX 3: كان بيعمل find للـ TAG_NAME "iframe" وبياخد أول iframe على الصفحة =====
        # الصح: نستخدم switch_to_dialog_frame الموجودة أصلاً وبتدور على الـ iframe الصح
        w.until(EC.element_to_be_clickable(self.get_profile_locator("add_client_btn"))).click()
        self.switch_to_dialog_frame(
            "iframe",  # fallback لو مفيش src محدد — الدالة ستجرب كل iframes
            self.get_profile_locator("client_petitioner_desc"),
            timeout=12
        )

        self.select_option_fuzzy(*self.get_profile_locator("client_petitioner_desc"), self.get_fixed("صفة_مقدم_الطلب"))
        self.select_option_fuzzy(*self.get_profile_locator("client_identity_type"), self.get_fixed("نوع_الهوية"))

        self.clear_and_type(*self.get_profile_locator("client_national_id"), self.get_fixed("الرقم_القومي"))
        self.clear_and_type(*self.get_profile_locator("client_first_name"), self.get_fixed("الاسم_الاول"))
        self.clear_and_type(*self.get_profile_locator("client_second_name"), self.get_fixed("الاسم_الثاني"))
        self.clear_and_type(*self.get_profile_locator("client_third_name"), self.get_fixed("الاسم_الثالث"))
        self.clear_and_type(*self.get_profile_locator("client_fourth_name"), self.get_fixed("الاسم_الرابع"))

        self.clear_and_type(*self.get_profile_locator("client_address"), self.get_fixed("العنوان"))
        self.clear_and_type(*self.get_profile_locator("client_email"), self.get_fixed("البريد_الالكتروني"))

        w.until(EC.element_to_be_clickable(self.get_profile_locator("dialog_save_btn"))).click()
        d.switch_to.default_content()

        doc_btn_locator = (
            By.CSS_SELECTOR,
            "button[aria-label='إضافة مستند'], "
            "button[title='إضافة مستند'], "
            "button[onclick*='add-attatchment']"
        )
        w.until(EC.element_to_be_clickable(doc_btn_locator)).click()
        self.switch_to_dialog_frame(
            self.get_profile_value("doc_dialog_frame_css"),
            self.get_profile_locator("attachment_type"),
            timeout=15
        )

        self.select_option_fuzzy(*self.get_profile_locator("attachment_type"), self.get_fixed("نوع_المستند"))
        w.until(EC.presence_of_element_located(self.get_profile_locator("attachment_input"))).send_keys(os.path.abspath(self.pdf_path.get()))
        w.until(EC.element_to_be_clickable(self.get_profile_locator("dialog_save_btn"))).click()
        d.switch_to.default_content()
        WebDriverWait(d, 12).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, self.get_profile_value("doc_dialog_frame_css")))
        )
        w.until(EC.element_to_be_clickable(self.get_profile_locator("delivery_receipt")))
        self.ui_sleep(0.5)

        self.select_option_fuzzy(*self.get_profile_locator("delivery_receipt"), self.get_fixed("طريقة_الإستلام"))
        self.select_option_fuzzy(*self.get_profile_locator("delivery_gov"), self.get_fixed("محافظة_التوصيل"))
        self.clear_and_type(*self.get_profile_locator("delivery_phone"), self.get_fixed("رقم_تليفون_للتواصل"))
        self.clear_and_type(*self.get_profile_locator("delivery_address"), self.get_fixed("عنوان_التوصيل"))
        w.until(EC.element_to_be_clickable(self.get_profile_locator("terms_checkbox"))).click()

        d.switch_to.default_content()
        self.mark_case_counted(idx)

    def mark_case_counted(self, case_idx):
        if case_idx in self.session_counted_case_indices:
            return
        self.session_counted_case_indices.add(case_idx)
        self.session_submitted_requests = len(self.session_counted_case_indices)

    def open_request_tab(self, request_url):
        d, w = self.driver, self.wait
        d.execute_script("window.open(arguments[0], '_blank');", request_url)
        d.switch_to.window(d.window_handles[-1])
        d.switch_to.default_content()
        w.until(EC.presence_of_element_located(self.get_profile_locator("request_ready")))

    def switch_to_case_tab(self, index):
        if index < 0 or index >= len(self.case_tabs):
            self.err("فهرس التبويب المطلوب غير صحيح")
        self.driver.switch_to.window(self.case_tabs[index])
        self.driver.switch_to.default_content()

    def focus_captcha_field(self):
        d = self.driver
        try:
            cap_elem = self.get_captcha_input_element(primary_timeout=2, fallback_timeout=4)
            d.execute_script("arguments[0].scrollIntoView({block:'center'});", cap_elem)
            try:
                cap_elem.click()
            except Exception:
                pass
        except Exception:
            pass

    def fill_captcha_in_current_tab(self, captcha_text):
        cap_elem = self.get_captcha_input_element(primary_timeout=2, fallback_timeout=6)
        try:
            cap_elem.clear()
        except Exception:
            pass
        cap_elem.send_keys(captcha_text)

    def click_first_visible(self, locators):
        d = self.driver
        for by, locator in locators:
            try:
                elements = d.find_elements(by, locator)
            except Exception:
                continue
            for elem in elements:
                try:
                    if not elem.is_displayed():
                        continue
                    d.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
                    try:
                        elem.click()
                    except Exception:
                        d.execute_script("arguments[0].click();", elem)
                    return True
                except Exception:
                    continue
        return False

    def click_submit_request_button(self):
        # ===== FIX 2: نضيف profile locator كأول محاولة — (By.ID, "cid-submit") =====
        # الكود القديم كان بيدور على id*='SUBMIT' وهو case-sensitive ومش بيطابق "cid-submit"
        profile_submit_locators = [self.get_profile_locator("submit_button")]
        if self.click_first_visible(profile_submit_locators):
            return True

        # fallback generic locators لو الـ profile locator فشل
        submit_locators = [
            (By.ID, "SUBMIT"),
            (By.CSS_SELECTOR, "button[id*='SUBMIT'],input[id*='SUBMIT'][type='button'],input[id*='SUBMIT'][type='submit']"),
            (By.CSS_SELECTOR, "button[title*='تقديم'],button[aria-label*='تقديم'],button[title*='إرسال'],button[aria-label*='إرسال']"),
            (By.CSS_SELECTOR, "button[onclick*='apex.submit'],a[onclick*='apex.submit'],button[onclick*='submit'],a[onclick*='submit']"),
            (By.XPATH, "//button[contains(normalize-space(.),'تقديم') or contains(normalize-space(.),'إرسال') or contains(normalize-space(.),'ارسال') or contains(normalize-space(.),'Submit')]"),
            (By.XPATH, "//a[contains(normalize-space(.),'تقديم') or contains(normalize-space(.),'إرسال') or contains(normalize-space(.),'ارسال') or contains(normalize-space(.),'Submit')]"),
        ]
        return self.click_first_visible(submit_locators)

    def click_optional_confirm(self):
        confirm_locators = [
            (By.XPATH, "//button[contains(normalize-space(.),'نعم') or contains(normalize-space(.),'تأكيد') or contains(normalize-space(.),'موافق') or normalize-space(.)='OK']"),
            (By.XPATH, "//a[contains(normalize-space(.),'نعم') or contains(normalize-space(.),'تأكيد') or contains(normalize-space(.),'موافق') or normalize-space(.)='OK']"),
            (By.CSS_SELECTOR, "button.ui-button--hot, .ui-dialog-buttonset button.t-Button--hot"),
        ]
        return self.click_first_visible(confirm_locators)

    def detect_submission_error(self):
        d = self.driver
        selectors = [
            ".t-Alert--danger",
            ".a-Notification--error",
            ".u-danger-text",
            ".t-Form-error",
        ]
        text = ""
        for sel in selectors:
            try:
                elems = d.find_elements(By.CSS_SELECTOR, sel)
                for e in elems:
                    if e.is_displayed():
                        txt = e.text.strip()
                        if txt:
                            text += f" {txt}"
            except Exception:
                pass

        text = text.strip()
        if not text:
            return None

        lowered = text.lower()
        keywords = ["خطأ", "غير صحيح", "كود التحقق", "captcha", "تحقق", "invalid"]
        if any(k in lowered for k in keywords):
            return text
        return None

    def submit_current_request(self, captcha_text):
        self.fill_captcha_in_current_tab(captcha_text)

        if not self.click_submit_request_button():
            return False, "تعذر العثور على زر تقديم الطلب."

        self.ui_sleep(0.3)
        self.click_optional_confirm()
        self.ui_sleep(1.0)
        err_text = self.detect_submission_error()
        if err_text:
            return False, err_text
        return True, ""

    def click_series_refresh(self):
        refresh_locators = []

        try:
            refresh_locators.append(self.get_profile_locator("series_refresh"))
        except Exception:
            pass

        # الأيقونة نفسها أو الزرار الأب كـ fallback لو locator الخدمة لم ينجح
        refresh_locators.extend([
            (By.XPATH, "//button[.//i[contains(@class,'cc-refresh')]]"),
            (By.XPATH, "//a[.//i[contains(@class,'cc-refresh')]]"),
            (By.CSS_SELECTOR, ".cc-refresh"),
        ])
        d = self.driver
        for by, locator in refresh_locators:
            try:
                elements = d.find_elements(by, locator)
                for elem in elements:
                    if elem.is_displayed():
                        d.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
                        try:
                            elem.click()
                        except Exception:
                            d.execute_script("arguments[0].click();", elem)
                        return True
            except Exception:
                continue

        try:
            icons = d.find_elements(By.CSS_SELECTOR, ".fas.fa-sync-alt.cc-refresh, .cc-refresh, i.cc-refresh")
            for icon in icons:
                if icon.is_displayed():
                    d.execute_script("arguments[0].parentElement.click();", icon)
                    return True
        except Exception:
            pass
        return False

    def read_series_display_text(self):
        d = self.driver
        elem = WebDriverWait(d, 4).until(
            EC.presence_of_element_located(self.get_profile_locator("series_display"))
        )
        txt = (elem.text or "").strip()
        if not txt:
            txt = (elem.get_attribute("value") or "").strip()
        return txt

    def fetch_request_number_current_tab(self, max_refresh_clicks=5):
        d = self.driver
        d.switch_to.default_content()

        for _ in range(max_refresh_clicks):
            self.click_series_refresh()
            self.ui_sleep(0.7)
            try:
                txt = self.read_series_display_text()
            except Exception:
                txt = ""

            if txt and re.fullmatch(r"\d+", txt):
                return txt

            if txt == self.get_profile_value("series_inprogress_text"):
                continue

        return ""

    def save_request_numbers_to_excel(self):
        path = self.excel_path.get()
        if not path or not os.path.isfile(path):
            self.err("مسار ملف Excel غير موجود لحفظ أرقام الطلبات.")

        try:
            wb = load_workbook(path)
            if "Cases_Data" not in wb.sheetnames:
                self.err("لم يتم العثور على الشيت Cases_Data في ملف Excel.")
            ws = wb["Cases_Data"]

            headers = {}
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=c).value
                if val is not None:
                    headers[str(val).strip()] = c

            req_col = headers.get("رقم_الطلب")
            if not req_col:
                req_col = ws.max_column + 1
                ws.cell(row=1, column=req_col, value="رقم_الطلب")

            for idx in range(len(self.cases)):
                v = self.cases.iloc[idx].get("رقم_الطلب", "")
                if pd.isna(v):
                    v = ""
                ws.cell(row=idx + 2, column=req_col, value=str(v).strip())

            wb.save(path)
        except PermissionError:
            self.err("تعذر حفظ ملف Excel. أغلق الملف من Excel ثم أعد المحاولة.")

    def save_single_inquiry_status_to_excel(self, row_idx, status_text):
        path = self.excel_path.get()
        if not path or not os.path.isfile(path):
            self.err("مسار ملف Excel غير موجود لحفظ حالة الطلب.")

        try:
            wb = load_workbook(path)
            if "Cases_Data" not in wb.sheetnames:
                self.err("لم يتم العثور على الشيت Cases_Data في ملف Excel.")
            ws = wb["Cases_Data"]

            headers = {}
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=c).value
                if val is not None:
                    headers[str(val).strip()] = c

            status_col = headers.get("حالة_الطلب")
            if not status_col:
                status_col = ws.max_column + 1
                ws.cell(row=1, column=status_col, value="حالة_الطلب")

            ws.cell(row=row_idx + 2, column=status_col, value=str(status_text).strip())
            wb.save(path)
        except PermissionError:
            self.err("تعذر حفظ ملف Excel. أغلق الملف من Excel ثم أعد المحاولة.")

    def activate_current_case_for_captcha(self):
        self.switch_to_case_tab(self.current_index)
        self.focus_captcha_field()
        if not self.step_frame.winfo_ismapped():
            self.step_frame.pack(fill=tk.X, pady=8)
        self.next_btn.config(state=tk.NORMAL)
        self.state = "waiting_captcha"
        global_idx = self.batch_case_indices[self.current_index] if self.batch_case_indices else self.current_index
        row = self.cases.iloc[global_idx]
        self.status_var.set(
            f"جاهز للكابتشا {global_idx+1}/{len(self.cases)} | رقم الطلب: {row['رقم_الطلب']}"
        )

    def normalize_request_number(self, value):
        if pd.isna(value):
            return ""
        text = str(value).strip()
        if re.fullmatch(r"\d+\.0", text):
            return text[:-2]
        return text

    def prepare_inquiry_current_tab(self, request_no, idx, total):
        d, w = self.driver, self.wait
        self.status_var.set(f"تجهيز استعلام {idx+1}/{total} | رقم الطلب: {request_no}")
        d.switch_to.default_content()
        field = w.until(EC.presence_of_element_located(self.get_profile_locator("inquiry_series_field")))
        try:
            field.clear()
        except Exception:
            pass
        field.send_keys(request_no)

    def activate_current_inquiry_for_captcha(self):
        self.switch_to_case_tab(self.current_index)
        self.focus_captcha_field()
        if not self.step_frame.winfo_ismapped():
            self.step_frame.pack(fill=tk.X, pady=8)
        self.next_btn.config(state=tk.NORMAL)
        self.state = "waiting_captcha"
        req_no = self.inquiry_rows[self.current_index]["request_no"]
        self.status_var.set(
            f"جاهز لكابتشا الاستعلام {self.current_index+1}/{len(self.inquiry_rows)} | رقم الطلب: {req_no}"
        )
        self.autofill_inquiry_captcha_from_ocr()

    def setup_ocr_engine(self):
        if Image is None or ImageOps is None or pytesseract is None:
            self.ocr_ready = False
            return

        current_cmd = getattr(getattr(pytesseract, "pytesseract", pytesseract), "tesseract_cmd", "") or ""
        if not current_cmd or (not os.path.isfile(current_cmd) and shutil.which(current_cmd) is None):
            common_paths = [
                r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                os.path.join(os.environ.get("LOCALAPPDATA", ""), "Programs", "Tesseract-OCR", "tesseract.exe"),
            ]
            for p in common_paths:
                if os.path.isfile(p):
                    pytesseract.pytesseract.tesseract_cmd = p
                    break

        cmd = getattr(getattr(pytesseract, "pytesseract", pytesseract), "tesseract_cmd", "") or ""
        self.ocr_ready = bool(cmd and (os.path.isfile(cmd) or shutil.which(cmd) is not None))

    def prepare_all_inquiry_tabs(self):
        d, w = self.driver, self.wait
        try:
            if "رقم_الطلب" not in self.cases.columns:
                self.err("لا يوجد عمود رقم_الطلب في Cases_Data.")

            self.inquiry_rows = []
            for i in range(len(self.cases)):
                req_no = self.normalize_request_number(self.cases.iloc[i].get("رقم_الطلب", ""))
                if req_no:
                    self.inquiry_rows.append({"row_idx": i, "request_no": req_no})

            total = len(self.inquiry_rows)
            if total == 0:
                self.err("لا توجد أرقام في عمود رقم_الطلب للاستعلام.")

            try:
                self.step_frame.pack_forget()
            except Exception:
                pass
            self.next_btn.config(state=tk.DISABLED)
            self.state = "ready"
            self.case_tabs = []
            self.current_index = 0

            d.switch_to.default_content()
            w.until(EC.presence_of_element_located(self.get_profile_locator("request_ready")))
            self.ensure_inquiry_status_column()
            inquiry_url = d.current_url

            for idx, item in enumerate(self.inquiry_rows):
                if idx == 0:
                    d.switch_to.default_content()
                else:
                    self.open_request_tab(inquiry_url)

                self.prepare_inquiry_current_tab(item["request_no"], idx, total)
                self.case_tabs.append(d.current_window_handle)

            self.current_index = 0
            self.activate_current_inquiry_for_captcha()
            messagebox.showinfo("جاهز", f"تم تجهيز {total} طلب للاستعلام.\nابدأ بإدخال كابتشا الطلب الأول.")
        except Exception as e:
            self.err(f"فشل تجهيز طلبات الاستعلام: {e}", raise_exc=False)

    def next_step_inquiry(self):
        w = self.wait
        if self.state != "waiting_captcha":
            return
        if not self.case_tabs or not self.inquiry_rows:
            self.err("لا يوجد طلبات استعلام مجهزة.")

        self.switch_to_case_tab(self.current_index)
        captcha_value = (self.captcha_code.get() or "").strip()
        if not captcha_value:
            self.err("ادخل الكابتشا أولًا.")

        captcha_elem = self.get_captcha_input_element(primary_timeout=2, fallback_timeout=6)
        try:
            captcha_elem.clear()
        except Exception:
            pass
        captcha_elem.send_keys(captcha_value)

        typed_value = (captcha_elem.get_attribute("value") or "").strip()
        if typed_value != captcha_value:
            try:
                captcha_elem.clear()
            except Exception:
                pass
            captcha_elem.send_keys(captcha_value)
            typed_value = (captcha_elem.get_attribute("value") or "").strip()
            if typed_value != captcha_value:
                self.status_var.set("تعذر كتابة الكابتشا بشكل صحيح. أعد كتابة الكابتشا لنفس الطلب.")
                self.focus_captcha_field()
                self.captcha_code.set("")
                self.autofill_inquiry_captcha_from_ocr()
                return

        submit_btn = w.until(EC.element_to_be_clickable(self.get_profile_locator("submit_button")))
        submit_btn.click()
        self.ui_sleep(0.3)

        err_text = self.detect_inquiry_captcha_error(timeout=3.5)
        if err_text:
            self.status_var.set(f"{err_text} | أعد إدخال الكابتشا لنفس الطلب.")
            self.focus_captcha_field()
            self.captcha_code.set("")
            self.autofill_inquiry_captcha_from_ocr()
            return

        inquiry_status = self.fetch_inquiry_status_current_tab(timeout=6.0)
        current_item = self.inquiry_rows[self.current_index]
        row_idx = current_item["row_idx"]

        # ===== FIX 6: ensure_inquiry_status_column كانت بتتنادى في كل iteration =====
        # نناديها مرة واحدة قبل اللوب في prepare_all_inquiry_tabs
        # هنا بس نكتب القيمة مباشرة
        self.cases.at[self.cases.index[row_idx], "حالة_الطلب"] = inquiry_status
        self.save_single_inquiry_status_to_excel(row_idx, inquiry_status)

        try:
            self.step_frame.pack_forget()
        except Exception:
            pass

        self.captcha_code.set("")
        self.state = "ready"
        self.next_btn.config(state=tk.DISABLED)

        if self.current_index + 1 >= len(self.case_tabs):
            self.hide_runtime_widgets_on_finish()
            self.status_var.set("تم تنفيذ الاستعلام وحفظ حالة_الطلب في Excel.")
            messagebox.showinfo("انتهى", "تم عرض حالة جميع الطلبات وحفظها في عمود حالة_الطلب.")
            return

        self.current_index += 1
        self.activate_current_inquiry_for_captcha()

    def read_inquiry_captcha_guess(self):
        d = self.driver
        try:
            cap_el = WebDriverWait(d, 3).until(
                EC.presence_of_element_located((By.ID, "P29_CAPTCHA"))
            )
        except Exception:
            return ""

        try:
            direct_candidates = [
                cap_el.text,
                cap_el.get_attribute("innerText"),
                cap_el.get_attribute("textContent"),
                cap_el.get_attribute("value"),
                cap_el.get_attribute("aria-label"),
                cap_el.get_attribute("title"),
                cap_el.get_attribute("alt"),
            ]
            for v in direct_candidates:
                txt = re.sub(r"[^A-Za-z0-9]", "", (v or ""))
                if 4 <= len(txt) <= 8:
                    return txt
        except Exception:
            pass

        if not self.ocr_ready:
            return ""

        try:
            png = cap_el.screenshot_as_png
            img = Image.open(io.BytesIO(png))
        except Exception:
            return ""

        try:
            base = img.convert("L")
            variants = [
                base,
                ImageOps.autocontrast(base),
                ImageOps.invert(ImageOps.autocontrast(base)),
                ImageOps.autocontrast(base).resize((base.width * 3, base.height * 3)),
            ]
        except Exception:
            return ""

        whitelist = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789"
        config = f"--oem 3 --psm 8 -c tessedit_char_whitelist={whitelist}"

        for candidate in variants:
            try:
                txt = pytesseract.image_to_string(candidate, config=config)
            except Exception:
                continue
            txt = re.sub(r"[^A-Za-z0-9]", "", (txt or ""))
            if 4 <= len(txt) <= 8:
                return txt

        return ""

    def autofill_inquiry_captcha_from_ocr(self):
        guess = self.read_inquiry_captcha_guess()
        if guess:
            self.captcha_code.set(guess)
            self.status_var.set(f"تم اقتراح الكابتشا: {guess} (راجعها ثم اضغط الخطوة التالية)")
            return

        if not self.ocr_ready and not self.ocr_warned:
            self.ocr_warned = True
            self.status_var.set("تعذر تفعيل OCR تلقائيًا. ثبّت Tesseract OCR أو اضبط مساره.")

    def detect_inquiry_captcha_error(self, timeout=3.5):
        d = self.driver
        end_time = time.time() + timeout
        keywords = ["كود الصورة غير صحيح", "captcha", "غير صحيح", "التحقق", "invalid"]

        while time.time() < end_time:
            try:
                err_elem = d.find_element(By.ID, "P29_CODE_error")
                err_text = (err_elem.text or "").strip()
                if err_elem.is_displayed() and err_text:
                    return err_text
            except Exception:
                pass

            try:
                msgs = d.find_elements(By.CSS_SELECTOR, ".t-Form-error, .a-Notification--error, .t-Alert--danger")
                for m in msgs:
                    if not m.is_displayed():
                        continue
                    t = (m.text or "").strip()
                    if t and any(k in t.lower() for k in [x.lower() for x in keywords]):
                        return t
            except Exception:
                pass

            self.ui_sleep(0.2)
        return ""

    def fetch_inquiry_status_current_tab(self, timeout=6.0):
        d = self.driver
        end_time = time.time() + timeout

        while time.time() < end_time:
            try:
                boxes = d.find_elements(By.CSS_SELECTOR, ".divBox")
                for box in boxes:
                    if not box.is_displayed():
                        continue
                    txt = (box.text or "").strip()
                    if not txt or "حالة الطلب" not in txt:
                        continue

                    parts = [p.strip() for p in txt.splitlines() if p.strip()]
                    if not parts:
                        continue

                    for i, part in enumerate(parts):
                        if "حالة الطلب" in part:
                            if i + 1 < len(parts):
                                return parts[i + 1]
                            cleaned = part.replace("حالة الطلب", "").strip(" :")
                            if cleaned:
                                return cleaned
                    return txt.replace("حالة الطلب", "").strip(" :")
            except Exception:
                pass
            self.ui_sleep(0.2)

        return ""

    def prepare_all_cases_tabs(self):
        self.prepare_cases_batch(0)

    def close_extra_case_tabs(self):
        d = self.driver
        if not self.case_tabs:
            return

        base_handle = self.case_tabs[0]
        for handle in list(self.case_tabs[1:]):
            try:
                d.switch_to.window(handle)
                d.close()
            except Exception:
                pass

        try:
            d.switch_to.window(base_handle)
            d.switch_to.default_content()
            self.case_tabs = [base_handle]
        except Exception:
            self.case_tabs = [d.current_window_handle]
            d.switch_to.default_content()

    def prepare_cases_batch(self, start_idx):
        d, w = self.driver, self.wait
        try:
            total = len(self.cases)
            if total == 0:
                self.err("Cases_Data فاضي")
            if start_idx >= total:
                self.err("لا يوجد طلبات متبقية للتجهيز.")

            try:
                self.step_frame.pack_forget()
            except Exception:
                pass
            self.next_btn.config(state=tk.DISABLED)
            self.state = "ready"
            self.current_index = 0

            if self.case_tabs:
                self.close_extra_case_tabs()

            d.switch_to.default_content()
            w.until(EC.presence_of_element_located(self.get_profile_locator("request_ready")))

            if not self.request_url:
                self.request_url = d.current_url

            end_idx = min(total, start_idx + max(1, int(self.max_open_tabs)))
            self.current_batch_start = start_idx
            self.batch_case_indices = list(range(start_idx, end_idx))
            self.case_tabs = []

            for offset, global_idx in enumerate(self.batch_case_indices):
                if offset == 0:
                    d.get(self.request_url)
                    d.switch_to.default_content()
                    w.until(EC.presence_of_element_located(self.get_profile_locator("request_ready")))
                else:
                    self.open_request_tab(self.request_url)

                row = self.cases.iloc[global_idx]
                self.prepare_case_in_current_tab(row, global_idx, total)
                self.case_tabs.append(d.current_window_handle)

            self.current_index = 0
            self.activate_current_case_for_captcha()
            prepared = len(self.batch_case_indices)
            self.status_var.set(
                f"تم تجهيز دفعة {prepared} طلب ({start_idx+1} - {end_idx}) من إجمالي {total}. جاهز للكابتشا."
            )
        except Exception as e:
            self.err(f"فشل تجهيز الطلبات في التبويبات: {e}", raise_exc=False)

    # ================= MANUAL RECOVERY =================
    def reload_current(self):
        try:
            if not self.driver or not self.wait:
                self.err("المتصفح غير مُهيأ بعد. ابدأ التشغيل أولاً.")
            if not self.case_tabs:
                self.err("لا يوجد طلبات مجهزة لإعادة تحميلها بعد.")

            self.status_var.set("جاري إعادة التحميل...")
            self.switch_to_case_tab(self.current_index)

            try:
                self.driver.switch_to.default_content()
            except Exception:
                pass

            self.driver.refresh()
            self.wait.until(EC.presence_of_element_located(self.get_profile_locator("request_ready")))

            try:
                self.step_frame.pack_forget()
            except Exception:
                pass

            if self.is_inquiry_service():
                req_no = self.inquiry_rows[self.current_index]["request_no"]
                self.prepare_inquiry_current_tab(req_no, self.current_index, len(self.inquiry_rows))
                self.activate_current_inquiry_for_captcha()
                self.status_var.set("تمت إعادة تجهيز طلب الاستعلام الحالي. أدخل الكابتشا مجددًا.")
                return

            # ===== حماية: تأكد إن batch_case_indices مش فاضي قبل الوصول =====
            if self.batch_case_indices:
                global_idx = self.batch_case_indices[self.current_index]
            else:
                global_idx = self.current_index

            row = self.cases.iloc[global_idx]
            self.prepare_case_in_current_tab(row, global_idx, len(self.cases))
            self.activate_current_case_for_captcha()
            self.status_var.set("تمت إعادة تجهيز الطلب الحالي. أدخل الكابتشا مجددًا.")
        except Exception as e:
            self.err(f"فشل إعادة التحميل: {e}", raise_exc=False)

    def next_step(self):
        try:
            if self.is_inquiry_service():
                self.next_step_inquiry()
                return

            if self.state != "waiting_captcha":
                return
            if not self.case_tabs:
                self.err("لا يوجد تبويبات طلبات للتنقل بينها.")

            self.switch_to_case_tab(self.current_index)
            captcha_value = (self.captcha_code.get() or "").strip()
            if not captcha_value:
                self.err("ادخل الكابتشا أولًا.")

            submitted_idx = self.current_index
            ok, err_text = self.submit_current_request(captcha_value)
            if not ok:
                self.status_var.set("فشل تحقق الكابتشا. جاري إعادة تحميل الطلب الحالي...")
                self.reload_current()
                return

            req_no = self.fetch_request_number_current_tab(max_refresh_clicks=8)
            if not req_no:
                self.status_var.set("لم يظهر رقم الطلب. جاري إعادة تحميل الطلب الحالي...")
                self.reload_current()
                return

            self.ensure_request_number_column()
            global_idx = self.batch_case_indices[submitted_idx] if self.batch_case_indices else submitted_idx
            self.cases.at[self.cases.index[global_idx], "رقم_الطلب"] = req_no

            try:
                self.step_frame.pack_forget()
            except Exception:
                pass

            self.captcha_code.set("")
            self.state = "ready"
            self.next_btn.config(state=tk.DISABLED)

            if self.current_index + 1 >= len(self.case_tabs):
                processed_until = self.current_batch_start + len(self.batch_case_indices)
                if processed_until < len(self.cases):
                    self.status_var.set(
                        f"انتهت الدفعة الحالية. تجهيز الدفعة التالية ({processed_until+1}/{len(self.cases)})..."
                    )
                    self.close_extra_case_tabs()
                    self.prepare_cases_batch(processed_until)
                    return

                self.status_var.set("تم تقديم كل الطلبات والتحقق من أرقامها...")
                self.save_request_numbers_to_excel()
                log_usage_online(
                    self.license_key.get(),
                    submitted_requests=self.session_submitted_requests
                )
                self.hide_runtime_widgets_on_finish()
                self.status_var.set("تم حفظ أرقام الطلبات في Excel.")
                messagebox.showinfo("انتهى", "تم تقديم جميع الطلبات وحفظ أرقام الطلبات في عمود رقم_الطلب.")
                return

            self.current_index += 1
            self.activate_current_case_for_captcha()

        except Exception as e:
            self.err(f"فشل الاستكمال بعد الكابتشا: {e}", raise_exc=False)


# ================= RUN =================
if __name__ == "__main__":
    App().mainloop()
